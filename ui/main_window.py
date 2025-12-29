"""
KitaplÄ±k UygulamasÄ± - Ana Pencere
=================================
UygulamanÄ±n ana arayÃ¼zÃ¼ burada tanÄ±mlanÄ±r.
"""

from PyQt6.QtWidgets import (
    QMainWindow,       # Ana pencere sÄ±nÄ±fÄ±
    QWidget,           # Genel konteyner
    QVBoxLayout,       # Dikey yerleÅŸim
    QHBoxLayout,       # Yatay yerleÅŸim
    QGridLayout,       # Grid yerleÅŸim
    QPushButton,       # Buton
    QTableWidget,      # Tablo
    QTableWidgetItem,  # Tablo hÃ¼cresi
    QHeaderView,       # Tablo baÅŸlÄ±k ayarlarÄ±
    QLabel,            # YazÄ± etiketi
    QLineEdit,         # Metin giriÅŸi
    QMessageBox,       # UyarÄ±/bilgi diyaloÄŸu
    QMenuBar,          # MenÃ¼ Ã§ubuÄŸu
    QMenu,             # MenÃ¼
    QSplitter,         # BÃ¶lÃ¼nebilir panel
    QStackedWidget,    # Sayfa deÄŸiÅŸtirici
    QScrollArea,       # KaydÄ±rÄ±labilir alan
    QFrame,            # Ã‡erÃ§eve
    QFileDialog,       # Dosya seÃ§ici
    QDialog,           # Dialog penceresi
    QGroupBox,         # Grup kutusu
    QCheckBox,         # Onay kutusu
    QComboBox,         # AÃ§Ä±lÄ±r liste
    QProgressDialog,   # Ä°lerleme dialog'u
    QListWidget,       # Liste widget
    QListWidgetItem,   # Liste Ã¶ÄŸesi
    QSpinBox,          # SayÄ± giriÅŸi
    QTextEdit,         # Ã‡ok satÄ±rlÄ± metin
    QFormLayout,       # Form yerleÅŸimi
)
from PyQt6.QtCore import Qt, QSize, QThread, pyqtSignal  # Hizalama sabitleri vs.
from PyQt6.QtGui import QFont, QAction, QPixmap  # Font ayarlarÄ±, menÃ¼ aksiyonlarÄ±, gÃ¶rsel

# Kendi modÃ¼llerimiz - bir Ã¼st klasÃ¶rden import
import sys
from pathlib import Path
sys.path.append(str(Path(__file__).parent.parent))
import database as db
from ui.book_dialog import SearchBookDialog, ManualBookDialog
from ui.themes import get_stylesheet, THEME_NAMES
from ui.shelf_panel import ShelfPanel
from ui.stats_dialog import StatsDialog
from ui.filter_bar import FilterBar


class MainWindow(QMainWindow):
    """
    Ana pencere sÄ±nÄ±fÄ±.
    
    QMainWindow'dan miras alÄ±yoruz (inheritance).
    Bu bize menÃ¼ Ã§ubuÄŸu, araÃ§ Ã§ubuÄŸu, durum Ã§ubuÄŸu gibi
    hazÄ±r Ã¶zellikler saÄŸlÄ±yor.
    """
    
    def __init__(self):
        """
        Pencere oluÅŸturulduÄŸunda Ã§alÄ±ÅŸÄ±r.
        super().__init__() ile Ã¼st sÄ±nÄ±fÄ±n __init__'ini Ã§aÄŸÄ±rÄ±yoruz.
        """
        super().__init__()
        
        # Mevcut tema
        self.current_theme = db.get_setting("theme", "dark")
        
        # SeÃ§ili raf (None = tÃ¼m kitaplar)
        self.current_shelf_id = None
        
        # GÃ¶rÃ¼nÃ¼m modu: "list" veya "grid"
        self.view_mode = db.get_setting("view_mode", "list")
        
        # Grid seÃ§im durumu
        self.selected_grid_cards = set()
        self.grid_cards = {}
        
        # Pencere ayarlarÄ±
        self.setWindowTitle("KitaplÄ±ÄŸÄ±m")
        self.setMinimumSize(1000, 700)
        
        # MenÃ¼ Ã§ubuÄŸunu oluÅŸtur
        self.setup_menu()
        
        # ArayÃ¼zÃ¼ oluÅŸtur
        self.setup_ui()
        
        # TemayÄ± uygula
        self.apply_theme(self.current_theme)
        
        # KitaplarÄ± yÃ¼kle
        self.load_books()
        
        # GÃ¶rÃ¼nÃ¼m butonlarÄ±nÄ± gÃ¼ncelle
        self.set_view_mode(self.view_mode)
    
    def setup_menu(self):
        """
        MenÃ¼ Ã§ubuÄŸunu oluÅŸturur.
        """
        menubar = self.menuBar()
        
        # === Dosya MenÃ¼sÃ¼ ===
        file_menu = menubar.addMenu("Dosya")
        
        # Kitap Ekle
        search_action = QAction("ğŸ” Ara ve Ekle...", self)
        search_action.setShortcut("Ctrl+N")
        search_action.triggered.connect(self.on_search_add_clicked)
        file_menu.addAction(search_action)
        
        manual_action = QAction("âœï¸ Manuel Ekle...", self)
        manual_action.setShortcut("Ctrl+Shift+N")
        manual_action.triggered.connect(self.on_manual_add_clicked)
        file_menu.addAction(manual_action)
        
        file_menu.addSeparator()
        
        # Ä°Ã§e/DÄ±ÅŸa Aktar
        import_action = QAction("ğŸ“¥ Ä°Ã§e Aktar...", self)
        import_action.triggered.connect(self.import_books)
        file_menu.addAction(import_action)
        
        export_menu = QMenu("ğŸ“¤ DÄ±ÅŸa Aktar", self)
        export_menu.addAction("CSV", lambda: self.export_books("csv"))
        export_menu.addAction("JSON", lambda: self.export_books("json"))
        export_menu.addAction("Excel", lambda: self.export_books("xlsx"))
        file_menu.addMenu(export_menu)
        
        file_menu.addSeparator()
        
        fetch_covers_action = QAction("ğŸ–¼ï¸ Eksik KapaklarÄ± Ä°ndir...", self)
        fetch_covers_action.triggered.connect(self.fetch_missing_covers)
        file_menu.addAction(fetch_covers_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction("Ã‡Ä±kÄ±ÅŸ", self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # === KitaplÄ±k MenÃ¼sÃ¼ ===
        library_menu = menubar.addMenu("KitaplÄ±k")
        
        reading_list_action = QAction("ğŸ“‹ Okuma Listesi", self)
        reading_list_action.setShortcut("Ctrl+L")
        reading_list_action.triggered.connect(self.show_reading_list)
        library_menu.addAction(reading_list_action)
        
        series_action = QAction("ğŸ“š Seriler", self)
        series_action.triggered.connect(self.show_series_dialog)
        library_menu.addAction(series_action)
        
        quotes_action = QAction("ğŸ’¬ AlÄ±ntÄ±lar", self)
        quotes_action.triggered.connect(self.show_all_quotes)
        library_menu.addAction(quotes_action)
        
        library_menu.addSeparator()
        
        stats_action = QAction("ğŸ“Š Ä°statistikler", self)
        stats_action.setShortcut("Ctrl+I")
        stats_action.triggered.connect(self.show_stats)
        library_menu.addAction(stats_action)
        
        goal_action = QAction("ğŸ¯ Okuma Hedefi", self)
        goal_action.triggered.connect(self.show_reading_goal)
        library_menu.addAction(goal_action)
        
        library_menu.addSeparator()
        
        ai_action = QAction("ğŸ¤– AI Asistan", self)
        ai_action.setShortcut("Ctrl+Shift+A")
        ai_action.triggered.connect(self.show_ai_assistant)
        library_menu.addAction(ai_action)
        
        # === GÃ¶rÃ¼nÃ¼m MenÃ¼sÃ¼ ===
        view_menu = menubar.addMenu("GÃ¶rÃ¼nÃ¼m")
        
        # Kenar Ã§ubuÄŸu
        self.sidebar_action = QAction("Kenar Ã‡ubuÄŸu", self)
        self.sidebar_action.setShortcut("Ctrl+B")
        self.sidebar_action.setCheckable(True)
        self.sidebar_action.setChecked(True)
        self.sidebar_action.triggered.connect(self.toggle_sidebar)
        view_menu.addAction(self.sidebar_action)
        
        view_menu.addSeparator()
        
        # GÃ¶rÃ¼nÃ¼m modu
        view_mode_menu = QMenu("GÃ¶rÃ¼nÃ¼m Modu", self)
        list_action = QAction("ğŸ“‹ Liste", self)
        list_action.triggered.connect(lambda: self.set_view_mode("list"))
        view_mode_menu.addAction(list_action)
        grid_action = QAction("ğŸ“· Grid (Kapaklar)", self)
        grid_action.triggered.connect(lambda: self.set_view_mode("grid"))
        view_mode_menu.addAction(grid_action)
        view_menu.addMenu(view_mode_menu)
        
        # SÃ¼tunlar
        columns_menu = QMenu("SÃ¼tunlar", self)
        self.column_actions = {}
        column_names = ["Kapak", "BaÅŸlÄ±k", "Yazar", "Sayfa", "Durum", "Puan"]
        for i, name in enumerate(column_names):
            action = QAction(name, self)
            action.setCheckable(True)
            action.setChecked(True)
            if i == 1:
                action.setEnabled(False)
            action.triggered.connect(lambda checked, col=i: self.toggle_column(col, checked))
            columns_menu.addAction(action)
            self.column_actions[i] = action
        columns_menu.addSeparator()
        columns_menu.addAction("TÃ¼mÃ¼nÃ¼ GÃ¶ster", self.show_all_columns_and_update_menu)
        view_menu.addMenu(columns_menu)
        
        view_menu.addSeparator()
        
        # Tema
        theme_menu = QMenu("ğŸ¨ Tema", self)
        theme_menu.addAction(THEME_NAMES["light"], lambda: self.apply_theme("light"))
        theme_menu.addAction(THEME_NAMES["dark"], lambda: self.apply_theme("dark"))
        view_menu.addMenu(theme_menu)
        
        # === YardÄ±m MenÃ¼sÃ¼ ===
        help_menu = menubar.addMenu("YardÄ±m")
        
        # BaÅŸlangÄ±Ã§ rehberi
        guide_action = QAction("ğŸ“– BaÅŸlangÄ±Ã§ Rehberi", self)
        guide_action.triggered.connect(self.show_guide)
        help_menu.addAction(guide_action)
        
        # Klavye kÄ±sayollarÄ±
        shortcuts_action = QAction("âŒ¨ï¸ Klavye KÄ±sayollarÄ±", self)
        shortcuts_action.setShortcut("F1")
        shortcuts_action.triggered.connect(self.show_shortcuts)
        help_menu.addAction(shortcuts_action)
        
        # Ã–zellikler
        features_action = QAction("âœ¨ Ã–zellikler", self)
        features_action.triggered.connect(self.show_features)
        help_menu.addAction(features_action)
        
        help_menu.addSeparator()
        
        # HakkÄ±nda
        about_action = QAction("â„¹ï¸ HakkÄ±nda", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
    
    def show_guide(self):
        """BaÅŸlangÄ±Ã§ rehberini gÃ¶sterir."""
        dialog = HelpDialog(self, "guide")
        dialog.exec()
    
    def show_shortcuts(self):
        """Klavye kÄ±sayollarÄ±nÄ± gÃ¶sterir."""
        dialog = HelpDialog(self, "shortcuts")
        dialog.exec()
    
    def show_features(self):
        """Ã–zellikleri gÃ¶sterir."""
        dialog = HelpDialog(self, "features")
        dialog.exec()
    
    def show_about(self):
        """HakkÄ±nda dialogunu gÃ¶sterir."""
        dialog = HelpDialog(self, "about")
        dialog.exec()
    
    def show_stats(self):
        """Ä°statistik penceresini aÃ§ar."""
        dialog = StatsDialog(self)
        dialog.exec()
    
    def show_reading_goal(self):
        """Okuma hedefi dialog'unu aÃ§ar."""
        dialog = ReadingGoalDialog(self)
        dialog.exec()
    
    def show_reading_list(self):
        """Okuma listesi dialog'unu aÃ§ar."""
        dialog = ReadingListDialog(self)
        dialog.exec()
    
    def show_all_quotes(self):
        """TÃ¼m alÄ±ntÄ±lar dialog'unu aÃ§ar."""
        dialog = AllQuotesDialog(self)
        dialog.exec()
    
    def show_series_dialog(self):
        """Kitap serileri dialog'unu aÃ§ar."""
        dialog = SeriesDialog(self)
        if dialog.exec():
            # Seri seÃ§ildiyse o serinin kitaplarÄ±nÄ± gÃ¶ster
            if dialog.selected_series:
                self.show_series_books(dialog.selected_series)
    
    def show_series_books(self, series_name: str):
        """Bir serinin kitaplarÄ±nÄ± gÃ¶sterir."""
        books = db.get_books_in_series(series_name)
        if books:
            self.current_shelf_id = None
            self.shelf_panel.shelf_list.setCurrentRow(0)
            self.search_input.clear()
            self.load_books(books)
            self.setWindowTitle(f"KitaplÄ±ÄŸÄ±m - ğŸ“š {series_name}")
    
    def show_ai_assistant(self):
        """AI Asistan dialog'unu aÃ§ar."""
        dialog = AIAssistantDialog(self)
        dialog.exec()
    
    def apply_theme(self, theme: str):
        """
        TemayÄ± uygular ve kaydeder.
        """
        self.current_theme = theme
        self.setStyleSheet(get_stylesheet(theme))
        
        # Tercihi kaydet
        db.set_setting("theme", theme)
    
    def setup_ui(self):
        """
        ArayÃ¼z bileÅŸenlerini oluÅŸturur ve yerleÅŸtirir.
        """
        # Ana widget - QMainWindow'un merkezine koyacaÄŸÄ±z
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Ana yerleÅŸim (dikey)
        # TÃ¼m bileÅŸenler Ã¼stten alta dizilecek
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(20, 20, 20, 20)  # Kenar boÅŸluklarÄ±
        main_layout.setSpacing(10)  # BileÅŸenler arasÄ± boÅŸluk
        
        # === ÃœST KISIM: BaÅŸlÄ±k, Filtreler ve Arama ===
        header_layout = QHBoxLayout()
        header_layout.setSpacing(15)
        
        # Sidebar toggle butonu
        self.sidebar_btn = QPushButton("â—€")
        self.sidebar_btn.setFixedSize(28, 28)
        self.sidebar_btn.setToolTip("Kenar Ã‡ubuÄŸunu Gizle/GÃ¶ster")
        self.sidebar_btn.clicked.connect(self.toggle_sidebar)
        header_layout.addWidget(self.sidebar_btn)
        
        # BaÅŸlÄ±k
        title_label = QLabel("ğŸ“š KitaplÄ±ÄŸÄ±m")
        title_font = QFont()
        title_font.setPointSize(20)
        title_font.setBold(True)
        title_label.setFont(title_font)
        header_layout.addWidget(title_label)
        
        # Filtreler (baÅŸlÄ±ktan sonra)
        self.filter_bar = FilterBar()
        self.filter_bar.filters_changed.connect(self.on_filters_changed)
        header_layout.addWidget(self.filter_bar)
        
        # Arama kutusu (en saÄŸda)
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("ğŸ” Ara...")
        self.search_input.setFixedWidth(200)
        self.search_input.textChanged.connect(self.on_search)
        header_layout.addWidget(self.search_input)
        
        # GÃ¶rÃ¼nÃ¼m butonlarÄ±
        self.list_view_btn = QPushButton("â˜°")
        self.list_view_btn.setFixedSize(32, 32)
        self.list_view_btn.setToolTip("Liste GÃ¶rÃ¼nÃ¼mÃ¼")
        self.list_view_btn.clicked.connect(lambda: self.set_view_mode("list"))
        header_layout.addWidget(self.list_view_btn)
        
        self.grid_view_btn = QPushButton("â–¦")
        self.grid_view_btn.setFixedSize(32, 32)
        self.grid_view_btn.setToolTip("Kapak GÃ¶rÃ¼nÃ¼mÃ¼")
        self.grid_view_btn.clicked.connect(lambda: self.set_view_mode("grid"))
        header_layout.addWidget(self.grid_view_btn)
        
        main_layout.addLayout(header_layout)
        
        # === ORTA KISIM: Raf Paneli + Kitap Tablosu ===
        # QSplitter ile yan yana, boyutu ayarlanabilir
        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Sol: Raf paneli
        self.shelf_panel = ShelfPanel()
        self.shelf_panel.shelf_selected.connect(self.on_shelf_selected)
        self.shelf_panel.all_books_selected.connect(self.on_all_books_selected)
        self.splitter.addWidget(self.shelf_panel)
        
        # SaÄŸ: GÃ¶rÃ¼nÃ¼mler ve butonlar
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(10)
        
        # Stacked widget - liste ve grid gÃ¶rÃ¼nÃ¼mleri iÃ§in
        self.view_stack = QStackedWidget()
        
        # === LÄ°STE GÃ–RÃœNÃœMÃœ (Tablo) ===
        self.books_table = QTableWidget()
        self.books_table.setColumnCount(6)
        self.books_table.setHorizontalHeaderLabels([
            "", "BaÅŸlÄ±k", "Yazar", "Sayfa", "Durum", "Puan"
        ])
        
        # SÃ¼tun geniÅŸlik ayarlarÄ±
        header = self.books_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)     # Kapak
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)   # BaÅŸlÄ±k
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)   # Yazar
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)     # Sayfa
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)     # Durum
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)     # Puan
        self.books_table.setColumnWidth(0, 60)
        self.books_table.setColumnWidth(3, 80)
        self.books_table.setColumnWidth(4, 100)
        self.books_table.setColumnWidth(5, 80)
        
        # SÃ¼tunlarÄ± sÃ¼rÃ¼kleyerek sÄ±ralama
        header.setSectionsMovable(True)
        header.setDragEnabled(True)
        header.setDragDropMode(QHeaderView.DragDropMode.InternalMove)
        
        # Header saÄŸ tÄ±k menÃ¼sÃ¼ (sÃ¼tun gizleme)
        header.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        header.customContextMenuRequested.connect(self.show_column_menu)
        
        # SatÄ±r seÃ§imi (Ã§oklu seÃ§im iÃ§in)
        self.books_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.books_table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        
        # Inline edit ve saÄŸ tÄ±k
        self.books_table.cellChanged.connect(self.on_cell_changed)
        self.books_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.books_table.customContextMenuRequested.connect(self.show_book_context_menu)
        
        self.view_stack.addWidget(self.books_table)
        
        # === GRID GÃ–RÃœNÃœMÃœ (Kapaklar) ===
        self.grid_scroll = QScrollArea()
        self.grid_scroll.setWidgetResizable(True)
        self.grid_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        
        self.grid_container = QWidget()
        self.grid_layout = QGridLayout(self.grid_container)
        self.grid_layout.setSpacing(20)
        self.grid_layout.setContentsMargins(10, 10, 10, 10)
        
        self.grid_scroll.setWidget(self.grid_container)
        self.view_stack.addWidget(self.grid_scroll)
        
        # GÃ¶rÃ¼nÃ¼m moduna gÃ¶re ayarla
        self.view_stack.setCurrentIndex(0 if self.view_mode == "list" else 1)
        
        right_layout.addWidget(self.view_stack)
        
        # Butonlar
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        search_add_btn = QPushButton("ğŸ” Ara ve Ekle")
        search_add_btn.setMinimumHeight(36)
        search_add_btn.setMinimumWidth(120)
        search_add_btn.clicked.connect(self.on_search_add_clicked)
        button_layout.addWidget(search_add_btn)
        
        manual_add_btn = QPushButton("âœï¸ Manuel Ekle")
        manual_add_btn.setMinimumHeight(36)
        manual_add_btn.setMinimumWidth(120)
        manual_add_btn.clicked.connect(self.on_manual_add_clicked)
        button_layout.addWidget(manual_add_btn)
        
        delete_button = QPushButton("ğŸ—‘ï¸ Sil")
        delete_button.setMinimumHeight(36)
        delete_button.setMinimumWidth(80)
        delete_button.clicked.connect(self.on_delete_book_clicked)
        button_layout.addWidget(delete_button)
        
        right_layout.addLayout(button_layout)
        
        self.splitter.addWidget(right_widget)
        
        # BaÅŸlangÄ±Ã§ boyutlarÄ± (sol panel dar, saÄŸ geniÅŸ)
        self.splitter.setSizes([200, 600])
        
        # Sidebar durumunu yÃ¼kle
        sidebar_hidden = db.get_setting("sidebar_hidden", "0") == "1"
        if sidebar_hidden:
            self.shelf_panel.hide()
            self.sidebar_btn.setText("â–¶")
        
        # SÃ¼tun ayarlarÄ±nÄ± yÃ¼kle
        self.load_column_settings()
        
        main_layout.addWidget(self.splitter)
    
    def load_books(self, books=None):
        """
        KitaplarÄ± tabloya yÃ¼kler.
        
        books parametresi verilmezse mevcut raf ve filtrelere gÃ¶re kitaplarÄ± Ã§eker.
        Arama sonuÃ§larÄ±nÄ± gÃ¶stermek iÃ§in parametre kullanÄ±lÄ±r.
        """
        if books is None:
            # Ã–nce raf filtresi
            if self.current_shelf_id is None:
                # Filtre Ã§ubuÄŸu varsa filtreleri uygula
                if hasattr(self, 'filter_bar') and self.filter_bar.has_active_filters():
                    filters = self.filter_bar.get_filters()
                    books = db.get_filtered_books(
                        status=filters["status"],
                        rating=filters["rating"],
                        year=filters["year"]
                    )
                else:
                    books = db.get_all_books()
            else:
                # Raf seÃ§iliyse, raftan kitaplarÄ± al
                books = db.get_books_in_shelf(self.current_shelf_id)
        
        # cellChanged sinyalini geÃ§ici olarak kapat (yÃ¼kleme sÄ±rasÄ±nda tetiklenmesin)
        self.books_table.blockSignals(True)
        
        # Tabloyu temizle
        self.books_table.setRowCount(0)
        
        # Thumbnail boyutu
        THUMB_HEIGHT = 50
        
        # Her kitap iÃ§in bir satÄ±r ekle
        for book in books:
            row = self.books_table.rowCount()
            self.books_table.insertRow(row)
            
            # SatÄ±r yÃ¼ksekliÄŸini ayarla
            self.books_table.setRowHeight(row, THUMB_HEIGHT + 10)
            
            # === KAPAK GÃ–RSELÄ° (dÃ¼zenlenemez) ===
            cover_label = QLabel()
            cover_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            if book["cover_path"]:
                pixmap = QPixmap(book["cover_path"])
                if not pixmap.isNull():
                    scaled = pixmap.scaledToHeight(
                        THUMB_HEIGHT, 
                        Qt.TransformationMode.SmoothTransformation
                    )
                    cover_label.setPixmap(scaled)
                else:
                    cover_label.setText("ğŸ“–")
            else:
                cover_label.setText("ğŸ“–")
            
            self.books_table.setCellWidget(row, 0, cover_label)
            
            # === DÃœZENLENEBÄ°LÄ°R SÃœTUNLAR ===
            # BaÅŸlÄ±k
            title_item = QTableWidgetItem(book["title"] or "")
            title_item.setData(Qt.ItemDataRole.UserRole, book["id"])  # ID'yi sakla
            self.books_table.setItem(row, 1, title_item)
            
            # Yazar
            self.books_table.setItem(row, 2, QTableWidgetItem(book["author"] or ""))
            
            # Sayfa sayÄ±sÄ±
            self.books_table.setItem(row, 3, QTableWidgetItem(
                str(book["page_count"]) if book["page_count"] else ""
            ))
            
            # Durum (dÃ¼zenlenebilir - metin olarak)
            status_map = {
                "unread": "ğŸ“• OkunmadÄ±",
                "to_read": "ğŸ“‹ OkuyacaÄŸÄ±m",
                "reading": "ğŸ“– Okunuyor",
                "read": "ğŸ“— Okundu",
                "wont_read": "ğŸš« OkumayacaÄŸÄ±m"
            }
            status_text = status_map.get(book["status"], book["status"])
            status_item = QTableWidgetItem(status_text)
            status_item.setData(Qt.ItemDataRole.UserRole + 1, book["status"])  # Orijinal deÄŸeri sakla
            self.books_table.setItem(row, 4, status_item)
            
            # Puan (dÃ¼zenlenebilir - metin olarak)
            rating = book["rating"]
            rating_text = "â­" * rating if rating else ""
            rating_item = QTableWidgetItem(rating_text)
            rating_item.setData(Qt.ItemDataRole.UserRole + 1, rating)  # Orijinal deÄŸeri sakla
            self.books_table.setItem(row, 5, rating_item)
        
        # Sinyalleri tekrar aÃ§
        self.books_table.blockSignals(False)
        
        # Grid gÃ¶rÃ¼nÃ¼mÃ¼nÃ¼ gÃ¼ncelle
        self.load_grid_view(books)
    
    def load_grid_view(self, books):
        """Grid gÃ¶rÃ¼nÃ¼mÃ¼nÃ¼ yÃ¼kler (kapak gÃ¶rselleri)."""
        # SeÃ§ili kartlarÄ± temizle
        self.selected_grid_cards = set()
        
        # Mevcut widget'larÄ± temizle
        while self.grid_layout.count():
            item = self.grid_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        if not books:
            empty_label = QLabel("Kitap bulunamadÄ±")
            empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.grid_layout.addWidget(empty_label, 0, 0)
            return
        
        # Grid boyutlarÄ±
        COVER_WIDTH = 130
        COVER_HEIGHT = 190
        CARD_WIDTH = COVER_WIDTH + 20
        
        # Dinamik sÃ¼tun sayÄ±sÄ± - mevcut geniÅŸliÄŸe gÃ¶re
        available_width = self.grid_scroll.viewport().width() - 40
        cols = max(2, available_width // (CARD_WIDTH + 20))
        
        # KartlarÄ± sakla
        self.grid_cards = {}
        
        # Grid'e kitaplarÄ± ekle
        for i, book in enumerate(books):
            row = i // cols
            col = i % cols
            
            # Kitap kartÄ±
            card = QFrame()
            card.setObjectName("bookCard")
            card.setFixedSize(CARD_WIDTH, COVER_HEIGHT + 50)
            card.setCursor(Qt.CursorShape.PointingHandCursor)
            card.setStyleSheet("""
                QFrame#bookCard {
                    background-color: transparent;
                    border-radius: 8px;
                    border: 2px solid transparent;
                }
                QFrame#bookCard:hover {
                    background-color: rgba(255, 255, 255, 0.05);
                }
            """)
            
            card_layout = QVBoxLayout(card)
            card_layout.setContentsMargins(5, 5, 5, 5)
            card_layout.setSpacing(6)
            
            # Kapak gÃ¶rseli
            cover_label = QLabel()
            cover_label.setFixedSize(COVER_WIDTH, COVER_HEIGHT)
            cover_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            cover_label.setStyleSheet("""
                background-color: #2D2D2D;
                border-radius: 4px;
            """)
            
            if book["cover_path"]:
                pixmap = QPixmap(book["cover_path"])
                if not pixmap.isNull():
                    scaled = pixmap.scaled(
                        COVER_WIDTH, COVER_HEIGHT,
                        Qt.AspectRatioMode.KeepAspectRatio,
                        Qt.TransformationMode.SmoothTransformation
                    )
                    cover_label.setPixmap(scaled)
                else:
                    cover_label.setText("ğŸ“–")
                    cover_label.setStyleSheet(cover_label.styleSheet() + "font-size: 48px;")
            else:
                cover_label.setText("ğŸ“–")
                cover_label.setStyleSheet(cover_label.styleSheet() + "font-size: 48px;")
            
            card_layout.addWidget(cover_label, alignment=Qt.AlignmentFlag.AlignCenter)
            
            # Kitap baÅŸlÄ±ÄŸÄ±
            title_label = QLabel(book["title"] or "")
            title_label.setWordWrap(True)
            title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            title_label.setMaximumWidth(COVER_WIDTH)
            title_label.setMaximumHeight(36)
            title_label.setStyleSheet("font-size: 11px;")
            card_layout.addWidget(title_label)
            
            # Kitap ID'sini sakla
            card.setProperty("book_id", book["id"])
            self.grid_cards[book["id"]] = card
            
            # TÄ±klama olaylarÄ±
            card.mousePressEvent = lambda event, bid=book["id"], c=card: self.on_grid_card_clicked(event, bid, c)
            card.mouseDoubleClickEvent = lambda event, bid=book["id"]: self.on_grid_card_double_clicked(event, bid)
            
            self.grid_layout.addWidget(card, row, col, alignment=Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        
        # SaÄŸa ve aÅŸaÄŸÄ±ya esnek boÅŸluk
        self.grid_layout.setColumnStretch(cols, 1)
        self.grid_layout.setRowStretch(row + 1, 1)
        
        # Kitap listesini sakla (resize iÃ§in)
        self._grid_books = books
    
    def resizeEvent(self, event):
        """Pencere boyutu deÄŸiÅŸtiÄŸinde grid'i yeniden Ã§iz."""
        super().resizeEvent(event)
        
        # Grid gÃ¶rÃ¼nÃ¼mÃ¼ndeyse ve kitaplar varsa
        if hasattr(self, '_grid_books') and self._grid_books:
            if self.view_mode == "grid":
                self.load_grid_view(self._grid_books)
    
    def on_grid_card_clicked(self, event, book_id, card):
        """Grid'de bir kitap kartÄ±na tek tÄ±klandÄ±ÄŸÄ±nda - seÃ§im."""
        if event.button() == Qt.MouseButton.LeftButton:
            # Ctrl tuÅŸu basÄ±lÄ± mÄ±?
            ctrl_pressed = event.modifiers() & Qt.KeyboardModifier.ControlModifier
            
            if ctrl_pressed:
                # Ã‡oklu seÃ§im - toggle
                if book_id in self.selected_grid_cards:
                    self.selected_grid_cards.remove(book_id)
                    self.update_card_selection(card, False)
                else:
                    self.selected_grid_cards.add(book_id)
                    self.update_card_selection(card, True)
            else:
                # Tek seÃ§im - diÄŸerlerini temizle
                self.clear_grid_selection()
                self.selected_grid_cards.add(book_id)
                self.update_card_selection(card, True)
                
        elif event.button() == Qt.MouseButton.RightButton:
            # SaÄŸ tÄ±k - menÃ¼ gÃ¶ster
            if book_id not in self.selected_grid_cards:
                # SeÃ§ili deÄŸilse Ã¶nce seÃ§
                self.clear_grid_selection()
                self.selected_grid_cards.add(book_id)
                self.update_card_selection(card, True)
            
            self.show_grid_context_menu(event.globalPosition().toPoint(), book_id)
    
    def on_grid_card_double_clicked(self, event, book_id):
        """Grid'de Ã§ift tÄ±klama - dÃ¼zenleme."""
        if event.button() == Qt.MouseButton.LeftButton:
            self.open_edit_dialog(book_id)
    
    def update_card_selection(self, card, selected: bool):
        """Kart seÃ§im stilini gÃ¼nceller."""
        if selected:
            card.setStyleSheet("""
                QFrame#bookCard {
                    background-color: rgba(0, 120, 212, 0.2);
                    border-radius: 8px;
                    border: 2px solid #0078D4;
                }
            """)
        else:
            card.setStyleSheet("""
                QFrame#bookCard {
                    background-color: transparent;
                    border-radius: 8px;
                    border: 2px solid transparent;
                }
                QFrame#bookCard:hover {
                    background-color: rgba(255, 255, 255, 0.05);
                }
            """)
    
    def clear_grid_selection(self):
        """TÃ¼m grid seÃ§imlerini temizler."""
        for book_id in list(self.selected_grid_cards):
            if book_id in self.grid_cards:
                self.update_card_selection(self.grid_cards[book_id], False)
        self.selected_grid_cards.clear()
    
    def show_grid_context_menu(self, position, book_id):
        """Grid gÃ¶rÃ¼nÃ¼mÃ¼nde saÄŸ tÄ±k menÃ¼sÃ¼."""
        # Ã‡oklu seÃ§im varsa toplu iÅŸlem menÃ¼sÃ¼ gÃ¶ster
        if len(self.selected_grid_cards) > 1:
            self.show_grid_multi_select_menu(position)
            return
        
        book = db.get_book_by_id(book_id)
        if not book:
            return
        
        menu = QMenu(self)
        
        # DÃ¼zenle
        edit_action = menu.addAction("âœï¸ DÃ¼zenle")
        edit_action.triggered.connect(lambda: self.open_edit_dialog(book_id))
        
        # Kopyala
        copy_action = menu.addAction("ğŸ“‹ Kopyala")
        copy_action.triggered.connect(lambda: self.copy_book(book_id))
        
        # AlÄ±ntÄ±lar
        quotes_action = menu.addAction("ğŸ’¬ AlÄ±ntÄ±lar")
        quotes_action.triggered.connect(lambda: self.show_quotes_dialog(book_id))
        
        menu.addSeparator()
        
        # Rafa ekle
        add_to_shelf_menu = QMenu("ğŸ“š Rafa Ekle", self)
        shelves = db.get_all_shelves()
        book_shelf_ids = [s["id"] for s in db.get_shelves_for_book(book_id)]
        
        for shelf in shelves:
            action_text = f"{shelf['icon']} {shelf['name']}"
            if shelf["id"] in book_shelf_ids:
                action_text = f"âœ“ {action_text}"
            
            action = add_to_shelf_menu.addAction(action_text)
            shelf_id = shelf["id"]
            if shelf_id in book_shelf_ids:
                action.triggered.connect(
                    lambda checked, sid=shelf_id, bid=book_id: self.remove_from_shelf(bid, sid)
                )
            else:
                action.triggered.connect(
                    lambda checked, sid=shelf_id, bid=book_id: self.add_to_shelf(bid, sid)
                )
        
        menu.addMenu(add_to_shelf_menu)
        
        menu.addSeparator()
        
        # Sil
        delete_action = menu.addAction("ğŸ—‘ï¸ Sil")
        delete_action.triggered.connect(lambda: self.delete_book(book_id, book["title"]))
        
        menu.exec(position)
    
    def show_grid_multi_select_menu(self, position):
        """Grid Ã§oklu seÃ§im iÃ§in saÄŸ tÄ±k menÃ¼sÃ¼."""
        book_ids = list(self.selected_grid_cards)
        
        menu = QMenu(self)
        menu.addAction(f"ğŸ“š {len(book_ids)} kitap seÃ§ili").setEnabled(False)
        menu.addSeparator()
        
        # Toplu dÃ¼zenleme
        bulk_edit_action = menu.addAction("âœï¸ Toplu DÃ¼zenle...")
        bulk_edit_action.triggered.connect(lambda: self.show_bulk_edit_dialog(book_ids))
        
        # Rafa ekle alt menÃ¼sÃ¼
        add_to_shelf_menu = QMenu("ğŸ“š Rafa Ekle", self)
        shelves = db.get_all_shelves()
        
        for shelf in shelves:
            action = add_to_shelf_menu.addAction(f"{shelf['icon']} {shelf['name']}")
            shelf_id = shelf["id"]
            action.triggered.connect(
                lambda checked, sid=shelf_id: self.bulk_add_to_shelf(book_ids, sid)
            )
        
        menu.addMenu(add_to_shelf_menu)
        
        menu.addSeparator()
        
        # Toplu silme
        delete_action = menu.addAction("ğŸ—‘ï¸ SeÃ§ilenleri Sil")
        delete_action.triggered.connect(lambda: self.bulk_delete_books(book_ids))
        
        menu.exec(position)
    
    def toggle_sidebar(self):
        """Kenar Ã§ubuÄŸunu gÃ¶ster/gizle."""
        if self.shelf_panel.isVisible():
            self.shelf_panel.hide()
            self.sidebar_btn.setText("â–¶")
            db.set_setting("sidebar_hidden", "1")
            if hasattr(self, 'sidebar_action'):
                self.sidebar_action.setChecked(False)
        else:
            self.shelf_panel.show()
            self.sidebar_btn.setText("â—€")
            db.set_setting("sidebar_hidden", "0")
            if hasattr(self, 'sidebar_action'):
                self.sidebar_action.setChecked(True)
    
    def show_column_menu(self, position):
        """SÃ¼tun gizleme/gÃ¶sterme menÃ¼sÃ¼."""
        menu = QMenu(self)
        
        column_names = ["Kapak", "BaÅŸlÄ±k", "Yazar", "Sayfa", "Durum", "Puan"]
        
        for i, name in enumerate(column_names):
            action = menu.addAction(name)
            action.setCheckable(True)
            action.setChecked(not self.books_table.isColumnHidden(i))
            
            # BaÅŸlÄ±k sÃ¼tunu her zaman gÃ¶rÃ¼nÃ¼r olmalÄ±
            if i == 1:
                action.setEnabled(False)
            else:
                action.triggered.connect(lambda checked, col=i: self.toggle_column(col, checked))
        
        menu.addSeparator()
        
        # TÃ¼mÃ¼nÃ¼ gÃ¶ster
        show_all = menu.addAction("TÃ¼mÃ¼nÃ¼ GÃ¶ster")
        show_all.triggered.connect(self.show_all_columns)
        
        menu.exec(self.books_table.horizontalHeader().mapToGlobal(position))
    
    def toggle_column(self, column: int, visible: bool):
        """SÃ¼tunu gizle/gÃ¶ster."""
        self.books_table.setColumnHidden(column, not visible)
        
        # MenÃ¼ action'Ä±nÄ± gÃ¼ncelle
        if hasattr(self, 'column_actions') and column in self.column_actions:
            self.column_actions[column].setChecked(visible)
        
        # AyarÄ± kaydet
        hidden_cols = []
        for i in range(self.books_table.columnCount()):
            if self.books_table.isColumnHidden(i):
                hidden_cols.append(str(i))
        db.set_setting("hidden_columns", ",".join(hidden_cols))
    
    def show_all_columns(self):
        """TÃ¼m sÃ¼tunlarÄ± gÃ¶ster."""
        for i in range(self.books_table.columnCount()):
            self.books_table.setColumnHidden(i, False)
        db.set_setting("hidden_columns", "")
    
    def show_all_columns_and_update_menu(self):
        """TÃ¼m sÃ¼tunlarÄ± gÃ¶ster ve menÃ¼yÃ¼ gÃ¼ncelle."""
        self.show_all_columns()
        if hasattr(self, 'column_actions'):
            for action in self.column_actions.values():
                action.setChecked(True)
    
    def load_column_settings(self):
        """SÃ¼tun ayarlarÄ±nÄ± yÃ¼kle."""
        hidden = db.get_setting("hidden_columns", "")
        if hidden:
            for col_str in hidden.split(","):
                try:
                    col = int(col_str)
                    if col != 1:  # BaÅŸlÄ±k her zaman gÃ¶rÃ¼nÃ¼r
                        self.books_table.setColumnHidden(col, True)
                        if hasattr(self, 'column_actions') and col in self.column_actions:
                            self.column_actions[col].setChecked(False)
                except ValueError:
                    pass
    
    def set_view_mode(self, mode: str):
        """GÃ¶rÃ¼nÃ¼m modunu deÄŸiÅŸtirir."""
        self.view_mode = mode
        db.set_setting("view_mode", mode)
        
        # Stack widget'Ä± gÃ¼ncelle
        self.view_stack.setCurrentIndex(0 if mode == "list" else 1)
        
        # Buton stillerini gÃ¼ncelle
        if mode == "list":
            self.list_view_btn.setStyleSheet("background-color: #0078D4;")
            self.grid_view_btn.setStyleSheet("")
        else:
            self.list_view_btn.setStyleSheet("")
            self.grid_view_btn.setStyleSheet("background-color: #0078D4;")
    
    def on_search(self, text):
        """
        Arama kutusuna yazÄ±ldÄ±ÄŸÄ±nda Ã§alÄ±ÅŸÄ±r.
        """
        if text.strip():
            books = db.search_books(text)
        else:
            books = db.get_all_books()
        
        self.load_books(books)
    
    def on_cell_changed(self, row, column):
        """
        HÃ¼cre dÃ¼zenlendiÄŸinde Ã§aÄŸrÄ±lÄ±r (inline edit).
        """
        # Kapak sÃ¼tunu dÃ¼zenlenemez (zaten widget)
        if column == 0:
            return
        
        # Kitap ID'sini al
        title_item = self.books_table.item(row, 1)
        if not title_item:
            return
        
        book_id = title_item.data(Qt.ItemDataRole.UserRole)
        if not book_id:
            return
        
        # Yeni deÄŸeri al
        item = self.books_table.item(row, column)
        new_value = item.text().strip()
        
        # SÃ¼tuna gÃ¶re gÃ¼ncelle
        if column == 1:  # BaÅŸlÄ±k
            if new_value:  # BaÅŸlÄ±k boÅŸ olamaz
                db.update_book(book_id, title=new_value)
            else:
                self.load_books()  # BoÅŸsa geri yÃ¼kle
                
        elif column == 2:  # Yazar
            db.update_book(book_id, author=new_value or None)
            
        elif column == 3:  # Sayfa
            try:
                page_count = int(new_value) if new_value else None
                db.update_book(book_id, page_count=page_count)
            except ValueError:
                self.load_books()  # GeÃ§ersiz sayÄ±, geri yÃ¼kle
                
        elif column == 4:  # Durum
            # Emoji veya metin olarak girilmiÅŸ olabilir
            status_input = new_value.lower()
            if "okunmadÄ±" in status_input or "unread" in status_input or "ğŸ“•" in new_value:
                db.update_book(book_id, status="unread")
            elif "okunuyor" in status_input or "reading" in status_input or "ğŸ“–" in new_value:
                db.update_book(book_id, status="reading")
            elif "okundu" in status_input or "read" in status_input or "ğŸ“—" in new_value:
                db.update_book(book_id, status="read")
            self.load_books()  # Durumu dÃ¼zgÃ¼n gÃ¶stermek iÃ§in yenile
            
        elif column == 5:  # Puan
            # YÄ±ldÄ±z sayÄ±sÄ±nÄ± say veya sayÄ± olarak al
            star_count = new_value.count("â­")
            if star_count > 0:
                rating = min(star_count, 5)
            else:
                try:
                    rating = int(new_value) if new_value else None
                    if rating and (rating < 1 or rating > 5):
                        rating = None
                except ValueError:
                    rating = None
            db.update_book(book_id, rating=rating)
            self.load_books()  # PuanÄ± yÄ±ldÄ±z olarak gÃ¶stermek iÃ§in yenile

    def on_search_add_clicked(self):
        """
        'Ara ve Ekle' butonuna tÄ±klanÄ±nca Ã§alÄ±ÅŸÄ±r.
        Online arama ile kitap ekler.
        """
        dialog = SearchBookDialog(self)
        
        if dialog.exec():
            data = dialog.get_data()
            if not data.get("title"):
                return
            
            # VeritabanÄ±na ekle (API'den gelen tÃ¼m alanlarla)
            book_id = db.add_book(
                title=data["title"],
                author=data.get("author"),
                isbn=data.get("isbn"),
                page_count=data.get("page_count"),
                publish_year=data.get("publish_year"),
                publisher=data.get("publisher"),
                cover_path=data.get("cover_path"),
                subtitle=data.get("subtitle"),
                description=data.get("description"),
                language=data.get("language"),
                categories=data.get("categories"),
            )
            
            # Tabloyu yenile
            self.load_books()
            self.shelf_panel.refresh()
            self.filter_bar.refresh_years()
    
    def on_manual_add_clicked(self):
        """
        'Manuel Ekle' butonuna tÄ±klanÄ±nca Ã§alÄ±ÅŸÄ±r.
        Manuel kitap giriÅŸi yapar.
        """
        dialog = ManualBookDialog(self)
        
        if dialog.exec():
            data = dialog.get_data()
            
            # VeritabanÄ±na ekle (temel + ek alanlar)
            book_id = db.add_book(
                title=data["title"],
                author=data["author"],
                isbn=data["isbn"],
                page_count=data["page_count"],
                publish_year=data["publish_year"],
                publisher=data["publisher"],
                cover_path=data["cover_path"],
                subtitle=data.get("subtitle"),
                description=data.get("description"),
                language=data.get("language"),
                categories=data.get("categories"),
                translator=data.get("translator"),
                original_title=data.get("original_title"),
                original_language=data.get("original_language"),
                series_name=data.get("series_name"),
                series_order=data.get("series_order"),
                format=data.get("format", "paperback"),
                location=data.get("location"),
                tags=data.get("tags"),
            )
            
            # Okuma takibi ve diÄŸer alanlarÄ± gÃ¼ncelle
            db.update_book(
                book_id,
                status=data.get("status", "unread"),
                current_page=data.get("current_page"),
                start_date=data.get("start_date"),
                finish_date=data.get("finish_date"),
                times_read=data.get("times_read"),
                rating=data.get("rating"),
                notes=data.get("notes"),
                review=data.get("review"),
                purchase_date=data.get("purchase_date"),
                purchase_place=data.get("purchase_place"),
                purchase_price=data.get("purchase_price"),
                currency=data.get("currency"),
                is_gift=data.get("is_gift"),
                gifted_by=data.get("gifted_by"),
                is_borrowed=data.get("is_borrowed"),
                borrowed_to=data.get("borrowed_to"),
                borrowed_date=data.get("borrowed_date"),
            )
            
            # Tabloyu yenile
            self.load_books()
            self.shelf_panel.refresh()
            self.filter_bar.refresh_years()
    
    def on_delete_book_clicked(self):
        """
        'Sil' butonuna tÄ±klanÄ±nca Ã§alÄ±ÅŸÄ±r.
        """
        # SeÃ§ili satÄ±rÄ± bul
        selected_items = self.books_table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen silmek iÃ§in bir kitap seÃ§in.")
            return
        
        row = selected_items[0].row()
        title_cell = self.books_table.item(row, 1)
        book_id = title_cell.data(Qt.ItemDataRole.UserRole)
        book_title = title_cell.text()
        
        self.delete_book(book_id, book_title)
    
    def on_shelf_selected(self, shelf_id: int):
        """Bir raf seÃ§ildiÄŸinde."""
        self.current_shelf_id = shelf_id
        self.search_input.clear()
        self.load_books()
    
    def on_all_books_selected(self):
        """'TÃ¼m Kitaplar' seÃ§ildiÄŸinde."""
        self.current_shelf_id = None
        self.search_input.clear()
        self.load_books()
    
    def on_filters_changed(self, filters: dict):
        """Filtreler deÄŸiÅŸtiÄŸinde."""
        # Arama kutusunu temizle (filtre ile arama karÄ±ÅŸmasÄ±n)
        self.search_input.clear()
        # Raf seÃ§imini "TÃ¼m Kitaplar"a al
        self.current_shelf_id = None
        self.shelf_panel.shelf_list.setCurrentRow(0)
        self.load_books()
    
    def show_book_context_menu(self, position):
        """Kitap tablosunda saÄŸ tÄ±k menÃ¼sÃ¼nÃ¼ gÃ¶sterir."""
        item = self.books_table.itemAt(position)
        if not item:
            return
        
        # SeÃ§ili satÄ±rlarÄ± al
        selected_rows = set()
        for item in self.books_table.selectedItems():
            selected_rows.add(item.row())
        
        selected_rows = list(selected_rows)
        
        # Tek kitap mÄ± Ã§oklu seÃ§im mi?
        if len(selected_rows) > 1:
            self.show_multi_select_menu(position, selected_rows)
            return
        
        row = item.row()
        title_cell = self.books_table.item(row, 1)
        if not title_cell:
            return
        
        book_id = title_cell.data(Qt.ItemDataRole.UserRole)
        book_title = title_cell.text()
        
        menu = QMenu(self)
        
        # DÃ¼zenle
        edit_action = menu.addAction("âœï¸ DÃ¼zenle")
        edit_action.triggered.connect(lambda: self.open_edit_dialog(book_id))
        
        # Kopyala
        copy_action = menu.addAction("ğŸ“‹ Kopyala")
        copy_action.triggered.connect(lambda: self.copy_book(book_id))
        
        # AlÄ±ntÄ±lar
        quotes_action = menu.addAction("ğŸ’¬ AlÄ±ntÄ±lar")
        quotes_action.triggered.connect(lambda: self.show_quotes_dialog(book_id))
        
        menu.addSeparator()
        
        # Rafa ekle alt menÃ¼sÃ¼
        add_to_shelf_menu = QMenu("ğŸ“š Rafa Ekle", self)
        
        # Mevcut raflarÄ± listele
        shelves = db.get_all_shelves()
        book_shelf_ids = [s["id"] for s in db.get_shelves_for_book(book_id)]
        
        for shelf in shelves:
            action_text = f"{shelf['icon']} {shelf['name']}"
            
            # Kitap zaten bu raftaysa iÅŸaretle
            if shelf["id"] in book_shelf_ids:
                action_text = f"âœ“ {action_text}"
            
            action = add_to_shelf_menu.addAction(action_text)
            
            # Lambda'da deÄŸiÅŸken yakalama sorunu iÃ§in default parametre
            shelf_id = shelf["id"]
            if shelf_id in book_shelf_ids:
                action.triggered.connect(
                    lambda checked, sid=shelf_id, bid=book_id: self.remove_from_shelf(bid, sid)
                )
            else:
                action.triggered.connect(
                    lambda checked, sid=shelf_id, bid=book_id: self.add_to_shelf(bid, sid)
                )
        
        menu.addMenu(add_to_shelf_menu)
        
        # EÄŸer bir rafta isek, "Bu raftan Ã§Ä±kar" seÃ§eneÄŸi
        if self.current_shelf_id is not None:
            menu.addSeparator()
            remove_action = menu.addAction("âŒ Bu Raftan Ã‡Ä±kar")
            remove_action.triggered.connect(
                lambda: self.remove_from_shelf(book_id, self.current_shelf_id)
            )
        
        menu.addSeparator()
        
        # Sil
        delete_action = menu.addAction("ğŸ—‘ï¸ Sil")
        delete_action.triggered.connect(lambda: self.delete_book(book_id, book_title))
        
        menu.exec(self.books_table.mapToGlobal(position))
    
    def show_multi_select_menu(self, position, selected_rows: list):
        """Ã‡oklu seÃ§im iÃ§in saÄŸ tÄ±k menÃ¼sÃ¼."""
        # Kitap ID'lerini topla
        book_ids = []
        for row in selected_rows:
            title_cell = self.books_table.item(row, 1)
            if title_cell:
                book_id = title_cell.data(Qt.ItemDataRole.UserRole)
                if book_id:
                    book_ids.append(book_id)
        
        if not book_ids:
            return
        
        menu = QMenu(self)
        menu.addAction(f"ğŸ“š {len(book_ids)} kitap seÃ§ili").setEnabled(False)
        menu.addSeparator()
        
        # Toplu dÃ¼zenleme
        bulk_edit_action = menu.addAction("âœï¸ Toplu DÃ¼zenle...")
        bulk_edit_action.triggered.connect(lambda: self.show_bulk_edit_dialog(book_ids))
        
        # Rafa ekle alt menÃ¼sÃ¼
        add_to_shelf_menu = QMenu("ğŸ“š Rafa Ekle", self)
        shelves = db.get_all_shelves()
        
        for shelf in shelves:
            action = add_to_shelf_menu.addAction(f"{shelf['icon']} {shelf['name']}")
            shelf_id = shelf["id"]
            action.triggered.connect(
                lambda checked, sid=shelf_id: self.bulk_add_to_shelf(book_ids, sid)
            )
        
        menu.addMenu(add_to_shelf_menu)
        
        menu.addSeparator()
        
        # Toplu silme
        delete_action = menu.addAction("ğŸ—‘ï¸ SeÃ§ilenleri Sil")
        delete_action.triggered.connect(lambda: self.bulk_delete_books(book_ids))
        
        menu.exec(self.books_table.mapToGlobal(position))
    
    def open_edit_dialog(self, book_id: int):
        """DÃ¼zenleme formunu aÃ§ar."""
        book = db.get_book_by_id(book_id)
        if book:
            dialog = ManualBookDialog(self, dict(book))
            
            if dialog.exec():
                data = dialog.get_data()
                
                # TÃ¼m alanlarÄ± gÃ¼ncelle
                db.update_book(
                    book_id,
                    # Temel
                    title=data["title"],
                    subtitle=data.get("subtitle"),
                    author=data.get("author"),
                    isbn=data.get("isbn"),
                    publisher=data.get("publisher"),
                    publish_year=data.get("publish_year"),
                    page_count=data.get("page_count"),
                    language=data.get("language"),
                    categories=data.get("categories"),
                    format=data.get("format"),
                    cover_path=data.get("cover_path"),
                    # Ã‡eviri & Seri
                    translator=data.get("translator"),
                    original_title=data.get("original_title"),
                    original_language=data.get("original_language"),
                    series_name=data.get("series_name"),
                    series_order=data.get("series_order"),
                    # Okuma takibi
                    status=data.get("status"),
                    current_page=data.get("current_page"),
                    start_date=data.get("start_date"),
                    finish_date=data.get("finish_date"),
                    times_read=data.get("times_read"),
                    rating=data.get("rating"),
                    # Notlar
                    notes=data.get("notes"),
                    review=data.get("review"),
                    description=data.get("description"),
                    # SatÄ±n alma
                    purchase_date=data.get("purchase_date"),
                    purchase_place=data.get("purchase_place"),
                    purchase_price=data.get("purchase_price"),
                    currency=data.get("currency"),
                    is_gift=data.get("is_gift"),
                    gifted_by=data.get("gifted_by"),
                    # Konum & Ã–dÃ¼nÃ§
                    location=data.get("location"),
                    is_borrowed=data.get("is_borrowed"),
                    borrowed_to=data.get("borrowed_to"),
                    borrowed_date=data.get("borrowed_date"),
                    # Etiketler
                    tags=data.get("tags"),
                )
                
                self.load_books()
    
    def copy_book(self, book_id: int):
        """KitabÄ± kopyalar."""
        new_id = db.copy_book(book_id)
        if new_id:
            QMessageBox.information(self, "BaÅŸarÄ±lÄ±", "Kitap kopyalandÄ±!")
            self.load_books()
            self.shelf_panel.refresh()
        else:
            QMessageBox.warning(self, "Hata", "Kitap kopyalanamadÄ±!")
    
    def show_quotes_dialog(self, book_id: int):
        """AlÄ±ntÄ±lar dialog'unu aÃ§ar."""
        book = db.get_book_by_id(book_id)
        if book:
            dialog = QuotesDialog(book_id, book["title"], self)
            dialog.exec()
    
    def show_bulk_edit_dialog(self, book_ids: list):
        """Toplu dÃ¼zenleme dialog'unu aÃ§ar."""
        dialog = BulkEditDialog(book_ids, self)
        if dialog.exec():
            self.load_books()
            self.shelf_panel.refresh()
    
    def bulk_delete_books(self, book_ids: list):
        """Birden fazla kitabÄ± siler."""
        reply = QMessageBox.question(
            self,
            "Toplu Silme",
            f"{len(book_ids)} kitabÄ± silmek istediÄŸinize emin misiniz?\n\nBu iÅŸlem geri alÄ±namaz!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            deleted = db.bulk_delete_books(book_ids)
            QMessageBox.information(self, "Silindi", f"{deleted} kitap silindi.")
            self.load_books()
            self.shelf_panel.refresh()
    
    def bulk_add_to_shelf(self, book_ids: list, shelf_id: int):
        """Birden fazla kitabÄ± rafa ekler."""
        added = db.bulk_add_to_shelf(book_ids, shelf_id)
        if added > 0:
            self.shelf_panel.refresh()
            QMessageBox.information(self, "Eklendi", f"{added} kitap rafa eklendi.")
    
    def delete_book(self, book_id: int, book_title: str):
        """KitabÄ± siler (onay ile)."""
        reply = QMessageBox.question(
            self,
            "Silme OnayÄ±",
            f'"{book_title}" kitabÄ±nÄ± silmek istediÄŸinize emin misiniz?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            db.delete_book(book_id)
            self.load_books()
            self.shelf_panel.refresh()
    
    def add_to_shelf(self, book_id: int, shelf_id: int):
        """KitabÄ± rafa ekler."""
        db.add_book_to_shelf(book_id, shelf_id)
        self.shelf_panel.refresh()
        
    def remove_from_shelf(self, book_id: int, shelf_id: int):
        """KitabÄ± raftan Ã§Ä±karÄ±r."""
        db.remove_book_from_shelf(book_id, shelf_id)
        self.shelf_panel.refresh()
        
        # EÄŸer o raftaysak, listeyi gÃ¼ncelle
        if self.current_shelf_id == shelf_id:
            self.load_books()
    
    def export_books(self, format: str):
        """
        KitaplarÄ± dÄ±ÅŸa aktarÄ±r.
        
        Args:
            format: "csv", "json" veya "xlsx"
        """
        # Dosya uzantÄ±larÄ± ve filtreleri
        filters = {
            "csv": "CSV DosyasÄ± (*.csv)",
            "json": "JSON DosyasÄ± (*.json)",
            "xlsx": "Excel DosyasÄ± (*.xlsx)",
        }
        
    def fetch_missing_covers(self):
        """KapaÄŸÄ± olmayan kitaplar iÃ§in online arama yaparak kapak indir."""
        from PyQt6.QtWidgets import QProgressDialog
        from PyQt6.QtCore import Qt, QCoreApplication
        from services.book_api import search_books, download_cover, cover_exists
        
        # KapaÄŸÄ± olmayan kitaplarÄ± bul (dosya gerÃ§ekten var mÄ± kontrol et)
        books = db.get_all_books()
        books_without_cover = [b for b in books if not cover_exists(b["cover_path"])]
        
        if not books_without_cover:
            QMessageBox.information(self, "Bilgi", "TÃ¼m kitaplarÄ±n kapaÄŸÄ± mevcut!")
            return
        
        reply = QMessageBox.question(
            self,
            "Kapak Ä°ndir",
            f"{len(books_without_cover)} kitabÄ±n kapaÄŸÄ± eksik.\n\n"
            f"Online arama yaparak kapaklarÄ± indirmek ister misiniz?\n"
            f"(Bu iÅŸlem birkaÃ§ dakika sÃ¼rebilir)",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Ä°lerleme dialog'u
        progress = QProgressDialog("Kapaklar indiriliyor...", "Ä°ptal", 0, len(books_without_cover), self)
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()
        
        downloaded = 0
        failed = 0
        
        for i, book in enumerate(books_without_cover):
            if progress.wasCanceled():
                break
            
            progress.setValue(i)
            progress.setLabelText(f"({i+1}/{len(books_without_cover)}) {book['title'][:40]}...")
            QCoreApplication.processEvents()
            
            # Arama yap
            search_query = book["title"]
            if book["author"]:
                search_query += " " + book["author"]
            
            try:
                results = search_books(search_query, "title")
                if results and results[0].cover_url:
                    cover_path = download_cover(
                        results[0].cover_url, 
                        results[0].isbn or str(book["id"])
                    )
                    if cover_path:
                        db.update_book(book["id"], cover_path=cover_path)
                        downloaded += 1
                    else:
                        failed += 1
                else:
                    failed += 1
            except Exception as e:
                print(f"Kapak indirme hatasÄ± ({book['title']}): {e}")
                failed += 1
        
        progress.setValue(len(books_without_cover))
        
        QMessageBox.information(
            self,
            "TamamlandÄ±",
            f"âœ… {downloaded} kapak indirildi\nâŒ {failed} kitap iÃ§in kapak bulunamadÄ±"
        )
        
        self.load_books()
    
    def export_books(self, format: str):
        """KitaplarÄ± dÄ±ÅŸa aktar."""
        filters = {
            "csv": "CSV DosyasÄ± (*.csv)",
            "json": "JSON DosyasÄ± (*.json)",
            "xlsx": "Excel DosyasÄ± (*.xlsx)"
        }
        
        # Dosya kaydetme dialogu
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "DÄ±ÅŸa Aktar",
            f"kitaplik.{format}",
            filters.get(format, "")
        )
        
        if not file_path:
            return
        
        # TÃ¼m kitaplarÄ± al
        books = db.get_all_books()
        
        # DÄ±ÅŸa aktarÄ±lacak alanlar
        export_fields = [
            "id", "title", "author", "isbn", "page_count", 
            "publish_year", "publisher", "status", "rating", 
            "notes", "start_date", "finish_date", "created_at"
        ]
        
        try:
            if format == "csv":
                self._export_csv(file_path, books, export_fields)
            elif format == "json":
                self._export_json(file_path, books, export_fields)
            elif format == "xlsx":
                self._export_xlsx(file_path, books, export_fields)
            
            QMessageBox.information(
                self,
                "BaÅŸarÄ±lÄ±",
                f"{len(books)} kitap dÄ±ÅŸa aktarÄ±ldÄ±:\n{file_path}"
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Hata",
                f"DÄ±ÅŸa aktarma hatasÄ±:\n{str(e)}"
            )
    
    def _export_csv(self, file_path: str, books, fields):
        """CSV olarak dÄ±ÅŸa aktar."""
        import csv
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            
            for book in books:
                row = {field: book[field] for field in fields}
                writer.writerow(row)
    
    def _export_json(self, file_path: str, books, fields):
        """JSON olarak dÄ±ÅŸa aktar."""
        import json
        
        data = []
        for book in books:
            row = {field: book[field] for field in fields}
            data.append(row)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    
    def _export_xlsx(self, file_path: str, books, fields):
        """Excel olarak dÄ±ÅŸa aktar."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise Exception("Excel iÃ§in 'openpyxl' paketi gerekli.\npip install openpyxl")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kitaplar"
        
        # BaÅŸlÄ±k satÄ±rÄ±
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
        header_text_color = Font(color="FFFFFF", bold=True)
        
        for col, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.font = header_text_color
            cell.fill = header_fill
        
        # Veri satÄ±rlarÄ±
        for row_idx, book in enumerate(books, 2):
            for col_idx, field in enumerate(fields, 1):
                ws.cell(row=row_idx, column=col_idx, value=book[field])
        
        # SÃ¼tun geniÅŸliklerini ayarla
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        wb.save(file_path)
    
    def import_books(self):
        """CSV veya Excel dosyasÄ±ndan kitap iÃ§e aktar."""
        from PyQt6.QtWidgets import QFileDialog
        
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Ä°Ã§e Aktar",
            "",
            "TÃ¼m Desteklenen (*.csv *.xlsx *.xls);;CSV (*.csv);;Excel (*.xlsx *.xls)"
        )
        
        if not file_path:
            return
        
        try:
            # DosyayÄ± oku
            if file_path.endswith('.csv'):
                import csv
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    # Ä°lk satÄ±rÄ± oku ve ayracÄ± tespit et
                    first_line = f.readline()
                    f.seek(0)
                    
                    # NoktalÄ± virgÃ¼l mÃ¼ virgÃ¼l mÃ¼?
                    if ';' in first_line and first_line.count(';') > first_line.count(','):
                        delimiter = ';'
                    else:
                        delimiter = ','
                    
                    reader = csv.DictReader(f, delimiter=delimiter)
                    rows = list(reader)
            else:
                # Excel
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active
                    
                    # Ä°lk satÄ±r baÅŸlÄ±k
                    headers = [cell.value for cell in ws[1]]
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                        rows.append(row_dict)
                except ImportError:
                    QMessageBox.critical(self, "Hata", "Excel iÃ§in openpyxl gerekli!\npip install openpyxl")
                    return
            
            if not rows:
                QMessageBox.warning(self, "UyarÄ±", "Dosyada veri bulunamadÄ±.")
                return
            
            # Ä°Ã§e aktarma dialog'unu aÃ§
            dialog = ImportDialog(rows, self)
            if dialog.exec():
                imported = dialog.imported_count
                QMessageBox.information(
                    self, 
                    "BaÅŸarÄ±lÄ±", 
                    f"{imported} kitap baÅŸarÄ±yla iÃ§e aktarÄ±ldÄ±."
                )
                self.load_books()
                self.shelf_panel.refresh()
                self.filter_bar.refresh_years()
        
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Ä°Ã§e aktarma hatasÄ±:\n{str(e)}")


class ImportDialog(QDialog):
    """Ä°Ã§e aktarma dialog'u - sÃ¼tun eÅŸleÅŸtirme ve Ã¶nizleme."""
    
    def __init__(self, rows: list, parent=None):
        super().__init__(parent)
        self.rows = rows
        self.imported_count = 0
        self.column_mappings = {}
        
        self.setWindowTitle("ğŸ“¥ Ä°Ã§e Aktar")
        self.setMinimumSize(800, 600)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Bilgi
        info = QLabel(f"ğŸ“„ {len(self.rows)} satÄ±r bulundu. SÃ¼tunlarÄ± eÅŸleÅŸtirin:")
        info.setStyleSheet("font-weight: bold; margin-bottom: 10px;")
        layout.addWidget(info)
        
        # SÃ¼tun eÅŸleÅŸtirme
        mapping_group = QGroupBox("SÃ¼tun EÅŸleÅŸtirme")
        mapping_layout = QGridLayout(mapping_group)
        
        file_columns = list(self.rows[0].keys()) if self.rows else []
        
        # KitaplÄ±k alanlarÄ±
        app_fields = [
            ("", "-- Atla --"),
            ("title", "BaÅŸlÄ±k *"),
            ("author", "Yazar"),
            ("publisher", "YayÄ±nevi"),
            ("isbn", "ISBN"),
            ("page_count", "Sayfa SayÄ±sÄ±"),
            ("publish_year", "YayÄ±n YÄ±lÄ±"),
            ("status", "Durum (Okundu/Okunuyor/OkunmadÄ±)"),
            ("rating", "Puan (1-5)"),
            ("shelf", "Raf"),
            ("categories", "Kategori"),
            ("notes", "Notlar"),
        ]
        
        self.combo_mappings = {}
        
        for i, file_col in enumerate(file_columns):
            row = i // 3
            col = (i % 3) * 2
            
            label = QLabel(f"{file_col}:")
            label.setStyleSheet("font-weight: bold;")
            mapping_layout.addWidget(label, row, col)
            
            combo = QComboBox()
            for value, display in app_fields:
                combo.addItem(display, value)
            
            # Otomatik eÅŸleÅŸtirme dene
            auto_match = self._guess_mapping(file_col)
            if auto_match:
                for j in range(combo.count()):
                    if combo.itemData(j) == auto_match:
                        combo.setCurrentIndex(j)
                        break
            
            mapping_layout.addWidget(combo, row, col + 1)
            self.combo_mappings[file_col] = combo
        
        layout.addWidget(mapping_group)
        
        # Ã–nizleme
        preview_group = QGroupBox("Ã–nizleme (ilk 5 satÄ±r)")
        preview_layout = QVBoxLayout(preview_group)
        
        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(len(file_columns))
        self.preview_table.setHorizontalHeaderLabels(file_columns)
        self.preview_table.setRowCount(min(5, len(self.rows)))
        
        for i, row in enumerate(self.rows[:5]):
            for j, col in enumerate(file_columns):
                item = QTableWidgetItem(str(row.get(col, "") or ""))
                self.preview_table.setItem(i, j, item)
        
        self.preview_table.resizeColumnsToContents()
        preview_layout.addWidget(self.preview_table)
        layout.addWidget(preview_group)
        
        # Online arama seÃ§eneÄŸi
        self.search_online = QCheckBox("ğŸ“¡ Online arama ile bilgileri tamamla (yavaÅŸ)")
        layout.addWidget(self.search_online)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("Ä°ptal")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        import_btn = QPushButton("ğŸ“¥ Ä°Ã§e Aktar")
        import_btn.clicked.connect(self.do_import)
        btn_layout.addWidget(import_btn)
        
        layout.addLayout(btn_layout)
    
    def _guess_mapping(self, col_name: str) -> str:
        """SÃ¼tun adÄ±na gÃ¶re otomatik eÅŸleÅŸtirme."""
        col_lower = col_name.lower().replace("_", " ").replace("-", " ")
        
        mappings = {
            "baÅŸlÄ±k": "title", "kitap": "title", "title": "title", "kitap adÄ±": "title",
            "yazar": "author", "author": "author",
            "yayÄ±nevi": "publisher", "publisher": "publisher", "yayinevi": "publisher",
            "isbn": "isbn",
            "sayfa": "page_count", "page": "page_count", "sayfa sayÄ±sÄ±": "page_count",
            "yÄ±l": "publish_year", "year": "publish_year", "yayÄ±n yÄ±lÄ±": "publish_year",
            "durum": "status", "status": "status", "okuma durumu": "status", "okuma": "status",
            "puan": "rating", "rating": "rating",
            "raf": "shelf", "shelf": "shelf",
            "kategori": "categories", "tÃ¼r": "categories", "category": "categories",
            "not": "notes", "notlar": "notes", "notes": "notes",
        }
        
        for key, value in mappings.items():
            if key in col_lower:
                return value
        return ""
    
    def do_import(self):
        """Ä°Ã§e aktarmayÄ± gerÃ§ekleÅŸtir."""
        from PyQt6.QtWidgets import QProgressDialog
        from PyQt6.QtCore import Qt, QCoreApplication
        
        # EÅŸleÅŸtirmeleri al
        mappings = {}
        for file_col, combo in self.combo_mappings.items():
            app_field = combo.currentData()
            if app_field:
                mappings[file_col] = app_field
        
        # BaÅŸlÄ±k eÅŸleÅŸtirmesi zorunlu
        if "title" not in mappings.values():
            QMessageBox.warning(self, "UyarÄ±", "BaÅŸlÄ±k sÃ¼tunu eÅŸleÅŸtirilmeli!")
            return
        
        # Online arama yapÄ±lacak mÄ±?
        do_online_search = self.search_online.isChecked()
        
        # API import
        if do_online_search:
            from services.book_api import search_books
        
        # Ä°lerleme dialog'u
        progress = QProgressDialog("Ä°Ã§e aktarÄ±lÄ±yor...", "Ä°ptal", 0, len(self.rows), self)
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()
        
        imported = 0
        shelf_cache = {}  # Raf adÄ± -> id
        
        for i, row in enumerate(self.rows):
            if progress.wasCanceled():
                break
            
            progress.setValue(i)
            QCoreApplication.processEvents()  # UI gÃ¼ncellemesi
            
            # SatÄ±rÄ± kitap verisine dÃ¶nÃ¼ÅŸtÃ¼r
            book_data = {}
            shelf_name = None
            status_value = "unread"
            
            for file_col, app_field in mappings.items():
                value = row.get(file_col)
                if value is None or str(value).strip() == "":
                    continue
                
                value = str(value).strip()
                
                if app_field == "shelf":
                    shelf_name = value
                elif app_field == "page_count":
                    try:
                        book_data[app_field] = int(float(value))
                    except:
                        pass
                elif app_field == "publish_year":
                    try:
                        book_data[app_field] = int(float(value))
                    except:
                        pass
                elif app_field == "rating":
                    try:
                        book_data[app_field] = min(5, max(1, int(float(value))))
                    except:
                        pass
                elif app_field == "status":
                    # Durum eÅŸleÅŸtirmesi
                    value_lower = value.lower()
                    if "okuduklarÄ±m" in value_lower or ("okundu" in value_lower and "okunmadÄ±" not in value_lower):
                        status_value = "read"
                    elif "okuyorum" in value_lower or "okuyor" in value_lower:
                        status_value = "reading"
                    else:
                        status_value = "unread"
                else:
                    book_data[app_field] = value
            
            # BaÅŸlÄ±k yoksa atla
            if not book_data.get("title"):
                continue
            
            progress.setLabelText(f"({i+1}/{len(self.rows)}) {book_data.get('title', '')[:40]}...")
            QCoreApplication.processEvents()
            
            # Online arama
            if do_online_search:
                search_query = book_data.get("title", "")
                if book_data.get("author"):
                    search_query += " " + book_data["author"]
                
                try:
                    results = search_books(search_query, "title")
                    if results:
                        # En iyi eÅŸleÅŸmeyi bul
                        best = results[0]
                        
                        # API'den gelen verileri ekle (mevcut olmayanlarÄ±)
                        if not book_data.get("isbn") and best.isbn:
                            book_data["isbn"] = best.isbn
                        if not book_data.get("page_count") and best.page_count:
                            book_data["page_count"] = best.page_count
                        if not book_data.get("publish_year") and best.publish_year:
                            book_data["publish_year"] = best.publish_year
                        if not book_data.get("publisher") and best.publisher:
                            book_data["publisher"] = best.publisher
                        if best.cover_url:
                            # KapaÄŸÄ± indir
                            from services.book_api import download_cover
                            cover_path = download_cover(best.cover_url, best.isbn or book_data.get("title", "book"))
                            if cover_path:
                                book_data["cover_path"] = cover_path
                        if best.description:
                            book_data["description"] = best.description[:1000]
                        if best.language:
                            book_data["language"] = best.language
                        if best.categories:
                            book_data["categories"] = best.categories
                except Exception as e:
                    print(f"API hatasÄ±: {e}")
            
            # KitabÄ± ekle
            book_id = db.add_book(**book_data)
            
            # Durumu gÃ¼ncelle
            db.update_book(book_id, status=status_value)
            
            imported += 1
            
            # Raf varsa ekle
            if shelf_name:
                if shelf_name not in shelf_cache:
                    # RafÄ± bul veya oluÅŸtur
                    shelves = db.get_all_shelves()
                    shelf_id = None
                    for shelf in shelves:
                        if shelf["name"].lower() == shelf_name.lower():
                            shelf_id = shelf["id"]
                            break
                    
                    if shelf_id is None:
                        shelf_id = db.add_shelf(shelf_name)
                    
                    shelf_cache[shelf_name] = shelf_id
                
                if shelf_cache.get(shelf_name):
                    db.add_book_to_shelf(book_id, shelf_cache[shelf_name])
        
        progress.setValue(len(self.rows))
        self.imported_count = imported
        self.accept()


# ============================================================
# ALINTILAR DIALOG'U
# ============================================================

class QuotesDialog(QDialog):
    """Kitap alÄ±ntÄ±larÄ± dialog'u."""
    
    def __init__(self, book_id: int, book_title: str, parent=None):
        super().__init__(parent)
        
        self.book_id = book_id
        self.book_title = book_title
        
        self.setWindowTitle(f"ğŸ’¬ AlÄ±ntÄ±lar - {book_title}")
        self.setMinimumSize(600, 500)
        self.setModal(True)
        
        self.setup_ui()
        self.load_quotes()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        
        # Yeni alÄ±ntÄ± ekleme alanÄ±
        add_group = QGroupBox("Yeni AlÄ±ntÄ± Ekle")
        add_layout = QVBoxLayout(add_group)
        
        self.quote_text = QTextEdit()
        self.quote_text.setPlaceholderText("AlÄ±ntÄ± metnini buraya yazÄ±n...")
        self.quote_text.setMaximumHeight(80)
        add_layout.addWidget(self.quote_text)
        
        details_layout = QHBoxLayout()
        
        self.page_input = QSpinBox()
        self.page_input.setRange(0, 99999)
        self.page_input.setSpecialValueText("Sayfa")
        self.page_input.setFixedWidth(80)
        details_layout.addWidget(QLabel("Sayfa:"))
        details_layout.addWidget(self.page_input)
        
        self.chapter_input = QLineEdit()
        self.chapter_input.setPlaceholderText("BÃ¶lÃ¼m (opsiyonel)")
        self.chapter_input.setFixedWidth(150)
        details_layout.addWidget(self.chapter_input)
        
        details_layout.addStretch()
        
        add_btn = QPushButton("â• Ekle")
        add_btn.clicked.connect(self.add_quote)
        details_layout.addWidget(add_btn)
        
        add_layout.addLayout(details_layout)
        layout.addWidget(add_group)
        
        # AlÄ±ntÄ± listesi
        self.quotes_list = QListWidget()
        self.quotes_list.setAlternatingRowColors(True)
        self.quotes_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.quotes_list.customContextMenuRequested.connect(self.show_quote_menu)
        layout.addWidget(self.quotes_list, stretch=1)
        
        # Kapat butonu
        close_btn = QPushButton("Kapat")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)
    
    def load_quotes(self):
        """AlÄ±ntÄ±larÄ± yÃ¼kler."""
        self.quotes_list.clear()
        
        quotes = db.get_quotes_by_book(self.book_id)
        
        for quote in quotes:
            text = quote["text"]
            if len(text) > 100:
                text = text[:100] + "..."
            
            info = []
            if quote["page_number"]:
                info.append(f"s.{quote['page_number']}")
            if quote["chapter"]:
                info.append(quote["chapter"])
            
            display = f'"{text}"'
            if info:
                display += f" ({', '.join(info)})"
            
            if quote["is_favorite"]:
                display = "â­ " + display
            
            item = QListWidgetItem(display)
            item.setData(Qt.ItemDataRole.UserRole, quote["id"])
            item.setData(Qt.ItemDataRole.UserRole + 1, quote["text"])
            self.quotes_list.addItem(item)
    
    def add_quote(self):
        """Yeni alÄ±ntÄ± ekler."""
        text = self.quote_text.toPlainText().strip()
        if not text:
            QMessageBox.warning(self, "UyarÄ±", "AlÄ±ntÄ± metni boÅŸ olamaz!")
            return
        
        page = self.page_input.value() or None
        chapter = self.chapter_input.text().strip() or None
        
        db.add_quote(self.book_id, text, page, chapter)
        
        # Formu temizle
        self.quote_text.clear()
        self.page_input.setValue(0)
        self.chapter_input.clear()
        
        self.load_quotes()
    
    def show_quote_menu(self, position):
        """AlÄ±ntÄ± saÄŸ tÄ±k menÃ¼sÃ¼."""
        item = self.quotes_list.itemAt(position)
        if not item:
            return
        
        quote_id = item.data(Qt.ItemDataRole.UserRole)
        full_text = item.data(Qt.ItemDataRole.UserRole + 1)
        
        menu = QMenu(self)
        
        # Tam metni gÃ¶ster
        view_action = menu.addAction("ğŸ‘ï¸ Tam Metni GÃ¶r")
        view_action.triggered.connect(lambda: QMessageBox.information(self, "AlÄ±ntÄ±", full_text))
        
        # Favori
        fav_action = menu.addAction("â­ Favori Yap/KaldÄ±r")
        fav_action.triggered.connect(lambda: self.toggle_favorite(quote_id))
        
        menu.addSeparator()
        
        # Sil
        delete_action = menu.addAction("ğŸ—‘ï¸ Sil")
        delete_action.triggered.connect(lambda: self.delete_quote(quote_id))
        
        menu.exec(self.quotes_list.mapToGlobal(position))
    
    def toggle_favorite(self, quote_id: int):
        """Favori durumunu deÄŸiÅŸtirir."""
        db.toggle_quote_favorite(quote_id)
        self.load_quotes()
    
    def delete_quote(self, quote_id: int):
        """AlÄ±ntÄ±yÄ± siler."""
        reply = QMessageBox.question(
            self, "Silme OnayÄ±",
            "Bu alÄ±ntÄ±yÄ± silmek istediÄŸinize emin misiniz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            db.delete_quote(quote_id)
            self.load_quotes()


# ============================================================
# TOPLU DÃœZENLEME DIALOG'U
# ============================================================

class BulkEditDialog(QDialog):
    """Birden fazla kitabÄ± dÃ¼zenleme dialog'u."""
    
    def __init__(self, book_ids: list, parent=None):
        super().__init__(parent)
        
        self.book_ids = book_ids
        
        self.setWindowTitle(f"âœï¸ Toplu DÃ¼zenle ({len(book_ids)} kitap)")
        self.setMinimumWidth(400)
        self.setModal(True)
        
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        info = QLabel(f"ğŸ“š {len(self.book_ids)} kitap seÃ§ili.\nSadece deÄŸiÅŸtirmek istediÄŸiniz alanlarÄ± doldurun.")
        info.setStyleSheet("color: #888;")
        layout.addWidget(info)
        
        form = QFormLayout()
        form.setSpacing(10)
        
        # Durum
        self.status_combo = QComboBox()
        self.status_combo.addItem("-- DeÄŸiÅŸtirme --", "")
        self.status_combo.addItem("ğŸ“• OkunmadÄ±", "unread")
        self.status_combo.addItem("ğŸ“‹ OkuyacaÄŸÄ±m", "to_read")
        self.status_combo.addItem("ğŸ“– Okunuyor", "reading")
        self.status_combo.addItem("ğŸ“— Okundu", "read")
        self.status_combo.addItem("ğŸš« OkumayacaÄŸÄ±m", "wont_read")
        form.addRow("Durum:", self.status_combo)
        
        # Puan
        self.rating_combo = QComboBox()
        self.rating_combo.addItem("-- DeÄŸiÅŸtirme --", 0)
        self.rating_combo.addItem("â­", 1)
        self.rating_combo.addItem("â­â­", 2)
        self.rating_combo.addItem("â­â­â­", 3)
        self.rating_combo.addItem("â­â­â­â­", 4)
        self.rating_combo.addItem("â­â­â­â­â­", 5)
        form.addRow("Puan:", self.rating_combo)
        
        # Dil
        self.language_input = QComboBox()
        self.language_input.setEditable(True)
        self.language_input.addItems(["", "tr", "en", "de", "fr", "es"])
        form.addRow("Dil:", self.language_input)
        
        # Konum
        self.location_input = QLineEdit()
        self.location_input.setPlaceholderText("Fiziksel konum")
        form.addRow("Konum:", self.location_input)
        
        # Etiketler
        self.tags_input = QLineEdit()
        self.tags_input.setPlaceholderText("Etiketler (virgÃ¼lle ayÄ±r)")
        form.addRow("Etiketler:", self.tags_input)
        
        layout.addLayout(form)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("Ä°ptal")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        save_btn = QPushButton("ğŸ’¾ Uygula")
        save_btn.clicked.connect(self.apply_changes)
        btn_layout.addWidget(save_btn)
        
        layout.addLayout(btn_layout)
    
    def apply_changes(self):
        """DeÄŸiÅŸiklikleri uygular."""
        updates = {}
        
        status = self.status_combo.currentData()
        if status:
            updates["status"] = status
        
        rating = self.rating_combo.currentData()
        if rating:
            updates["rating"] = rating
        
        language = self.language_input.currentText().strip()
        if language:
            updates["language"] = language
        
        location = self.location_input.text().strip()
        if location:
            updates["location"] = location
        
        tags = self.tags_input.text().strip()
        if tags:
            updates["tags"] = tags
        
        if not updates:
            QMessageBox.warning(self, "UyarÄ±", "HiÃ§bir alan deÄŸiÅŸtirilmedi!")
            return
        
        updated = db.bulk_update_books(self.book_ids, **updates)
        QMessageBox.information(self, "BaÅŸarÄ±lÄ±", f"{updated} kitap gÃ¼ncellendi.")
        self.accept()


# ============================================================
# OKUMA HEDEFÄ° DIALOG'U
# ============================================================

class ReadingGoalDialog(QDialog):
    """Okuma hedefi belirleme dialog'u."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        
        self.setWindowTitle("ğŸ¯ Okuma Hedefi")
        self.setMinimumWidth(400)
        self.setModal(True)
        
        self.setup_ui()
        self.load_goals()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        # Yeni hedef ekleme
        add_group = QGroupBox("Yeni Hedef Belirle")
        add_layout = QHBoxLayout(add_group)
        
        from datetime import datetime
        current_year = datetime.now().year
        
        self.year_input = QSpinBox()
        self.year_input.setRange(2000, 2100)
        self.year_input.setValue(current_year)
        add_layout.addWidget(QLabel("YÄ±l:"))
        add_layout.addWidget(self.year_input)
        
        self.target_input = QSpinBox()
        self.target_input.setRange(1, 1000)
        self.target_input.setValue(12)
        add_layout.addWidget(QLabel("Hedef:"))
        add_layout.addWidget(self.target_input)
        add_layout.addWidget(QLabel("kitap"))
        
        add_btn = QPushButton("â• Ekle")
        add_btn.clicked.connect(self.add_goal)
        add_layout.addWidget(add_btn)
        
        layout.addWidget(add_group)
        
        # Mevcut hedefler
        self.goals_list = QListWidget()
        self.goals_list.setAlternatingRowColors(True)
        layout.addWidget(self.goals_list, stretch=1)
        
        # Kapat
        close_btn = QPushButton("Kapat")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)
    
    def load_goals(self):
        """Hedefleri yÃ¼kler."""
        self.goals_list.clear()
        
        goals = db.get_all_reading_goals()
        
        for goal in goals:
            progress_bar = f"[{'â–ˆ' * int(goal['progress'] / 10)}{'â–‘' * (10 - int(goal['progress'] / 10))}]"
            
            text = f"{goal['year']}: {goal['completed']}/{goal['target_books']} kitap ({goal['progress']}%) {progress_bar}"
            
            item = QListWidgetItem(text)
            item.setData(Qt.ItemDataRole.UserRole, goal["year"])
            
            # Hedef tamamlandÄ±ysa yeÅŸil
            if goal['completed'] >= goal['target_books']:
                item.setForeground(Qt.GlobalColor.green)
            
            self.goals_list.addItem(item)
    
    def add_goal(self):
        """Yeni hedef ekler."""
        year = self.year_input.value()
        target = self.target_input.value()
        
        db.set_reading_goal(year, target)
        self.load_goals()
        QMessageBox.information(self, "BaÅŸarÄ±lÄ±", f"{year} iÃ§in {target} kitap hedefi belirlendi!")


# ============================================================
# TÃœM ALINTILAR DIALOG'U
# ============================================================

class AllQuotesDialog(QDialog):
    """TÃ¼m kitaplardan alÄ±ntÄ±larÄ± gÃ¶steren dialog."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        
        self.setWindowTitle("ğŸ’¬ TÃ¼m AlÄ±ntÄ±lar")
        self.setMinimumSize(700, 500)
        self.setModal(True)
        
        self.setup_ui()
        self.load_quotes()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        
        # Filtre
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("ğŸ”"))
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("AlÄ±ntÄ±larda ara...")
        self.search_input.textChanged.connect(self.filter_quotes)
        filter_layout.addWidget(self.search_input)
        
        self.fav_only = QCheckBox("Sadece Favoriler")
        self.fav_only.toggled.connect(self.load_quotes)
        filter_layout.addWidget(self.fav_only)
        
        layout.addLayout(filter_layout)
        
        # AlÄ±ntÄ± listesi
        self.quotes_list = QListWidget()
        self.quotes_list.setAlternatingRowColors(True)
        self.quotes_list.setWordWrap(True)
        self.quotes_list.itemDoubleClicked.connect(self.show_full_quote)
        layout.addWidget(self.quotes_list, stretch=1)
        
        # Kapat
        close_btn = QPushButton("Kapat")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)
    
    def load_quotes(self):
        """AlÄ±ntÄ±larÄ± yÃ¼kler."""
        self.quotes_list.clear()
        
        quotes = db.get_all_quotes()
        
        for quote in quotes:
            # Favori filtresi
            if self.fav_only.isChecked() and not quote["is_favorite"]:
                continue
            
            text = quote["text"]
            if len(text) > 150:
                text = text[:150] + "..."
            
            display = f'"{text}"\nğŸ“– {quote["book_title"]} - {quote["book_author"] or "Bilinmeyen"}'
            
            if quote["page_number"]:
                display += f" (s.{quote['page_number']})"
            
            if quote["is_favorite"]:
                display = "â­ " + display
            
            item = QListWidgetItem(display)
            item.setData(Qt.ItemDataRole.UserRole, quote["id"])
            item.setData(Qt.ItemDataRole.UserRole + 1, quote["text"])
            item.setData(Qt.ItemDataRole.UserRole + 2, quote["book_title"])
            self.quotes_list.addItem(item)
    
    def filter_quotes(self, text: str):
        """AlÄ±ntÄ±larÄ± filtreler."""
        text = text.lower()
        
        for i in range(self.quotes_list.count()):
            item = self.quotes_list.item(i)
            full_text = item.data(Qt.ItemDataRole.UserRole + 1).lower()
            book_title = item.data(Qt.ItemDataRole.UserRole + 2).lower()
            
            visible = text in full_text or text in book_title
            item.setHidden(not visible)
    
    def show_full_quote(self, item):
        """Tam alÄ±ntÄ±yÄ± gÃ¶sterir."""
        full_text = item.data(Qt.ItemDataRole.UserRole + 1)
        book_title = item.data(Qt.ItemDataRole.UserRole + 2)
        
        QMessageBox.information(self, f"ğŸ“– {book_title}", full_text)


# ============================================================
# OKUMA LÄ°STESÄ° DIALOG'U
# ============================================================

class ReadingListDialog(QDialog):
    """Okuma listesi - sÄ±ralÄ±, tahmini sÃ¼re ile."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        
        self.setWindowTitle("ğŸ“‹ Okuma Listesi")
        self.setMinimumSize(800, 600)
        self.setModal(True)
        
        # Okuma hÄ±zÄ± ayarlarÄ±
        self.pages_per_day = 30  # GÃ¼nlÃ¼k sayfa
        
        self.setup_ui()
        self.load_reading_list()
    
    def setup_ui(self):
        layout = QHBoxLayout(self)
        layout.setSpacing(15)
        
        # Sol: Okuma listesi
        left_layout = QVBoxLayout()
        
        left_layout.addWidget(QLabel("ğŸ“‹ Okuma SÄ±ram:"))
        
        self.reading_list = QListWidget()
        self.reading_list.setAlternatingRowColors(True)
        self.reading_list.setDragDropMode(QListWidget.DragDropMode.InternalMove)
        self.reading_list.currentItemChanged.connect(self.on_book_selected)
        self.reading_list.model().rowsMoved.connect(self.on_list_reordered)
        left_layout.addWidget(self.reading_list, stretch=1)
        
        # SÄ±ralama butonlarÄ±
        order_layout = QHBoxLayout()
        
        up_btn = QPushButton("â¬†ï¸ YukarÄ±")
        up_btn.clicked.connect(lambda: self.move_book("up"))
        order_layout.addWidget(up_btn)
        
        down_btn = QPushButton("â¬‡ï¸ AÅŸaÄŸÄ±")
        down_btn.clicked.connect(lambda: self.move_book("down"))
        order_layout.addWidget(down_btn)
        
        remove_btn = QPushButton("âŒ Listeden Ã‡Ä±kar")
        remove_btn.clicked.connect(self.remove_from_list)
        order_layout.addWidget(remove_btn)
        
        left_layout.addLayout(order_layout)
        
        layout.addLayout(left_layout, stretch=2)
        
        # SaÄŸ: Detaylar ve ayarlar
        right_layout = QVBoxLayout()
        
        # Okuma hÄ±zÄ± ayarlarÄ±
        speed_group = QGroupBox("â±ï¸ Okuma HÄ±zÄ± AyarlarÄ±")
        speed_layout = QFormLayout(speed_group)
        speed_layout.setSpacing(10)
        
        self.pages_per_day_input = QSpinBox()
        self.pages_per_day_input.setRange(1, 500)
        self.pages_per_day_input.setValue(self.pages_per_day)
        self.pages_per_day_input.valueChanged.connect(self.on_speed_changed)
        speed_layout.addRow("GÃ¼nlÃ¼k sayfa:", self.pages_per_day_input)
        
        self.minutes_per_page = QSpinBox()
        self.minutes_per_page.setRange(1, 30)
        self.minutes_per_page.setValue(2)
        self.minutes_per_page.valueChanged.connect(self.on_speed_changed)
        speed_layout.addRow("Sayfa baÅŸÄ± dakika:", self.minutes_per_page)
        
        right_layout.addWidget(speed_group)
        
        # SeÃ§ili kitap bilgisi
        book_group = QGroupBox("ğŸ“– SeÃ§ili Kitap")
        book_layout = QVBoxLayout(book_group)
        
        self.book_title_label = QLabel("Kitap seÃ§in")
        self.book_title_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        self.book_title_label.setWordWrap(True)
        book_layout.addWidget(self.book_title_label)
        
        self.book_author_label = QLabel("")
        self.book_author_label.setStyleSheet("color: #888;")
        book_layout.addWidget(self.book_author_label)
        
        self.book_pages_label = QLabel("")
        book_layout.addWidget(self.book_pages_label)
        
        self.book_estimate_label = QLabel("")
        self.book_estimate_label.setStyleSheet("font-size: 13px; color: #4EC9B0;")
        book_layout.addWidget(self.book_estimate_label)
        
        right_layout.addWidget(book_group)
        
        # Toplam istatistik
        stats_group = QGroupBox("ğŸ“Š Toplam")
        stats_layout = QVBoxLayout(stats_group)
        
        self.total_books_label = QLabel("Kitap: 0")
        stats_layout.addWidget(self.total_books_label)
        
        self.total_pages_label = QLabel("Sayfa: 0")
        stats_layout.addWidget(self.total_pages_label)
        
        self.total_time_label = QLabel("Tahmini sÃ¼re: -")
        self.total_time_label.setStyleSheet("font-weight: bold; color: #CCA700;")
        stats_layout.addWidget(self.total_time_label)
        
        self.finish_date_label = QLabel("BitiÅŸ tarihi: -")
        self.finish_date_label.setStyleSheet("color: #4EC9B0;")
        stats_layout.addWidget(self.finish_date_label)
        
        right_layout.addWidget(stats_group)
        
        # Kitap ekle
        add_group = QGroupBox("â• Listeye Ekle")
        add_layout = QVBoxLayout(add_group)
        
        self.candidates_combo = QComboBox()
        self.candidates_combo.setPlaceholderText("OkunmamÄ±ÅŸ kitaplardan seÃ§...")
        add_layout.addWidget(self.candidates_combo)
        
        add_btn = QPushButton("â• Listeye Ekle")
        add_btn.clicked.connect(self.add_to_list)
        add_layout.addWidget(add_btn)
        
        right_layout.addWidget(add_group)
        
        right_layout.addStretch()
        
        # Kapat
        close_btn = QPushButton("Kapat")
        close_btn.clicked.connect(self.accept)
        right_layout.addWidget(close_btn)
        
        layout.addLayout(right_layout, stretch=1)
    
    def load_reading_list(self):
        """Okuma listesini yÃ¼kler."""
        self.reading_list.clear()
        
        books = db.get_reading_list()
        
        for i, book in enumerate(books, start=1):
            pages = book["page_count"] or 0
            estimate = self.calculate_days(pages)
            
            text = f"{i}. {book['title']}"
            if book["author"]:
                text += f" - {book['author']}"
            text += f"\n   ğŸ“„ {pages} sayfa â€¢ â±ï¸ ~{estimate} gÃ¼n"
            
            item = QListWidgetItem(text)
            item.setData(Qt.ItemDataRole.UserRole, book["id"])
            item.setData(Qt.ItemDataRole.UserRole + 1, dict(book))
            self.reading_list.addItem(item)
        
        # Aday kitaplarÄ± yÃ¼kle
        self.load_candidates()
        
        # Ä°statistikleri gÃ¼ncelle
        self.update_stats()
    
    def load_candidates(self):
        """Listeye eklenebilecek kitaplarÄ± yÃ¼kler."""
        self.candidates_combo.clear()
        self.candidates_combo.addItem("OkunmamÄ±ÅŸ kitaplardan seÃ§...", None)
        
        books = db.get_books_to_read_candidates()
        
        for book in books:
            text = book["title"]
            if book["author"]:
                text += f" - {book['author']}"
            self.candidates_combo.addItem(text, book["id"])
    
    def calculate_days(self, pages: int) -> int:
        """Sayfa sayÄ±sÄ±ndan gÃ¼n hesaplar."""
        if pages <= 0 or self.pages_per_day <= 0:
            return 0
        return max(1, round(pages / self.pages_per_day))
    
    def calculate_hours(self, pages: int) -> float:
        """Sayfa sayÄ±sÄ±ndan saat hesaplar."""
        minutes = pages * self.minutes_per_page.value()
        return round(minutes / 60, 1)
    
    def on_book_selected(self, current, previous):
        """Kitap seÃ§ildiÄŸinde detaylarÄ± gÃ¶ster."""
        if not current:
            return
        
        book = current.data(Qt.ItemDataRole.UserRole + 1)
        if not book:
            return
        
        self.book_title_label.setText(book.get("title", ""))
        self.book_author_label.setText(book.get("author", "") or "")
        
        pages = book.get("page_count") or 0
        self.book_pages_label.setText(f"ğŸ“„ {pages} sayfa")
        
        days = self.calculate_days(pages)
        hours = self.calculate_hours(pages)
        self.book_estimate_label.setText(f"â±ï¸ ~{days} gÃ¼n ({hours} saat)")
    
    def on_speed_changed(self):
        """Okuma hÄ±zÄ± deÄŸiÅŸtiÄŸinde."""
        self.pages_per_day = self.pages_per_day_input.value()
        self.load_reading_list()
    
    def on_list_reordered(self):
        """Liste sÃ¼rÃ¼kle-bÄ±rak ile yeniden sÄ±ralandÄ±ÄŸÄ±nda."""
        book_ids = []
        for i in range(self.reading_list.count()):
            item = self.reading_list.item(i)
            book_ids.append(item.data(Qt.ItemDataRole.UserRole))
        
        db.reorder_reading_list(book_ids)
        self.load_reading_list()
    
    def move_book(self, direction: str):
        """KitabÄ± yukarÄ±/aÅŸaÄŸÄ± taÅŸÄ±r."""
        current = self.reading_list.currentItem()
        if not current:
            return
        
        book_id = current.data(Qt.ItemDataRole.UserRole)
        db.move_in_reading_list(book_id, direction)
        self.load_reading_list()
    
    def add_to_list(self):
        """SeÃ§ili kitabÄ± listeye ekler."""
        book_id = self.candidates_combo.currentData()
        if not book_id:
            return
        
        db.add_to_reading_list(book_id)
        self.load_reading_list()
    
    def remove_from_list(self):
        """SeÃ§ili kitabÄ± listeden Ã§Ä±karÄ±r."""
        current = self.reading_list.currentItem()
        if not current:
            return
        
        book_id = current.data(Qt.ItemDataRole.UserRole)
        db.remove_from_reading_list(book_id)
        self.load_reading_list()
    
    def update_stats(self):
        """Toplam istatistikleri gÃ¼nceller."""
        books = db.get_reading_list()
        
        total_books = len(books)
        total_pages = sum(b["page_count"] or 0 for b in books)
        total_days = self.calculate_days(total_pages)
        total_hours = self.calculate_hours(total_pages)
        
        self.total_books_label.setText(f"ğŸ“š {total_books} kitap")
        self.total_pages_label.setText(f"ğŸ“„ {total_pages} sayfa")
        self.total_time_label.setText(f"â±ï¸ ~{total_days} gÃ¼n ({total_hours} saat)")
        
        # BitiÅŸ tarihi
        from datetime import datetime, timedelta
        finish_date = datetime.now() + timedelta(days=total_days)
        self.finish_date_label.setText(f"ğŸ“… Tahmini bitiÅŸ: {finish_date.strftime('%d %B %Y')}")


# ============================================================
# SERÄ°LER DIALOG'U
# ============================================================

class SeriesDialog(QDialog):
    """Kitap serileri dialog'u."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        
        self.selected_series = None
        
        self.setWindowTitle("ğŸ“š Kitap Serileri")
        self.setMinimumSize(700, 500)
        self.setModal(True)
        
        self.setup_ui()
        self.load_series()
    
    def setup_ui(self):
        layout = QHBoxLayout(self)
        layout.setSpacing(15)
        
        # Sol: Seri listesi
        left_layout = QVBoxLayout()
        
        left_layout.addWidget(QLabel("ğŸ“š Seriler"))
        
        self.series_list = QListWidget()
        self.series_list.setAlternatingRowColors(True)
        self.series_list.currentItemChanged.connect(self.on_series_selected)
        self.series_list.itemDoubleClicked.connect(self.on_series_double_clicked)
        left_layout.addWidget(self.series_list)
        
        layout.addLayout(left_layout, stretch=1)
        
        # SaÄŸ: Seri detaylarÄ±
        right_layout = QVBoxLayout()
        
        # Seri baÅŸlÄ±ÄŸÄ±
        self.series_title = QLabel("Seri seÃ§in")
        self.series_title.setStyleSheet("font-size: 18px; font-weight: bold; padding: 10px;")
        right_layout.addWidget(self.series_title)
        
        # Ä°statistikler
        stats_frame = QFrame()
        stats_frame.setStyleSheet("background-color: #2D2D2D; border-radius: 8px; padding: 10px;")
        stats_layout = QGridLayout(stats_frame)
        
        self.stat_total = QLabel("Toplam: -")
        self.stat_read = QLabel("Okunan: -")
        self.stat_reading = QLabel("Okuyor: -")
        self.stat_unread = QLabel("OkunmadÄ±: -")
        self.stat_pages = QLabel("Toplam sayfa: -")
        self.stat_rating = QLabel("Ortalama puan: -")
        
        stats_layout.addWidget(self.stat_total, 0, 0)
        stats_layout.addWidget(self.stat_read, 0, 1)
        stats_layout.addWidget(self.stat_reading, 1, 0)
        stats_layout.addWidget(self.stat_unread, 1, 1)
        stats_layout.addWidget(self.stat_pages, 2, 0)
        stats_layout.addWidget(self.stat_rating, 2, 1)
        
        right_layout.addWidget(stats_frame)
        
        # Ä°lerleme Ã§ubuÄŸu
        self.progress_label = QLabel("")
        self.progress_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progress_label.setStyleSheet("font-size: 14px; padding: 10px;")
        right_layout.addWidget(self.progress_label)
        
        # Serideki kitaplar
        right_layout.addWidget(QLabel("ğŸ“– Serideki Kitaplar:"))
        
        self.books_list = QListWidget()
        self.books_list.setAlternatingRowColors(True)
        right_layout.addWidget(self.books_list, stretch=1)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        
        show_btn = QPushButton("ğŸ“– Seriyi GÃ¶ster")
        show_btn.clicked.connect(self.show_series)
        btn_layout.addWidget(show_btn)
        
        close_btn = QPushButton("Kapat")
        close_btn.clicked.connect(self.reject)
        btn_layout.addWidget(close_btn)
        
        right_layout.addLayout(btn_layout)
        
        layout.addLayout(right_layout, stretch=2)
    
    def load_series(self):
        """Serileri yÃ¼kler."""
        self.series_list.clear()
        
        series_list = db.get_all_series()
        
        if not series_list:
            item = QListWidgetItem("HenÃ¼z seri yok")
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsSelectable)
            self.series_list.addItem(item)
            return
        
        for series in series_list:
            # Durum ikonu
            if series["is_complete"]:
                icon = "âœ…"  # TamamÄ± okundu
            elif series["read_count"] > 0:
                icon = "ğŸ“–"  # Bir kÄ±smÄ± okundu
            else:
                icon = "ğŸ“š"  # HiÃ§ okunmadÄ±
            
            text = f"{icon} {series['name']} ({series['book_count']} kitap)"
            
            item = QListWidgetItem(text)
            item.setData(Qt.ItemDataRole.UserRole, series["name"])
            self.series_list.addItem(item)
    
    def on_series_selected(self, current, previous):
        """Seri seÃ§ildiÄŸinde detaylarÄ± gÃ¶ster."""
        if not current:
            return
        
        series_name = current.data(Qt.ItemDataRole.UserRole)
        if not series_name:
            return
        
        # Ä°statistikleri al
        stats = db.get_series_stats(series_name)
        if not stats:
            return
        
        # BaÅŸlÄ±k
        self.series_title.setText(f"ğŸ“š {series_name}")
        
        # Ä°statistikler
        self.stat_total.setText(f"ğŸ“š Toplam: {stats['total']} kitap")
        self.stat_read.setText(f"âœ… Okunan: {stats['read_count']}")
        self.stat_reading.setText(f"ğŸ“– Okuyor: {stats['reading_count']}")
        self.stat_unread.setText(f"ğŸ“• OkunmadÄ±: {stats['unread_count']}")
        self.stat_pages.setText(f"ğŸ“„ Sayfa: {stats['total_pages'] or 0}")
        
        if stats['avg_rating']:
            self.stat_rating.setText(f"â­ Puan: {stats['avg_rating']}")
        else:
            self.stat_rating.setText("â­ Puan: -")
        
        # Ä°lerleme
        progress = stats['progress']
        bar = "â–ˆ" * (progress // 10) + "â–‘" * (10 - progress // 10)
        
        if stats['is_complete']:
            self.progress_label.setText(f"ğŸ‰ Seri tamamlandÄ±! [{bar}] {progress}%")
            self.progress_label.setStyleSheet("color: #4EC9B0; font-size: 14px; padding: 10px;")
        else:
            self.progress_label.setText(f"Ä°lerleme: [{bar}] {progress}%")
            self.progress_label.setStyleSheet("color: #CCA700; font-size: 14px; padding: 10px;")
        
        # KitaplarÄ± listele
        self.books_list.clear()
        books = db.get_books_in_series(series_name)
        
        for book in books:
            # Durum ikonu
            status_icons = {"read": "âœ…", "reading": "ğŸ“–", "unread": "ğŸ“•"}
            icon = status_icons.get(book["status"], "ğŸ“•")
            
            # SÄ±ra numarasÄ±
            order = f"#{book['series_order']}" if book["series_order"] else ""
            
            # Puan
            rating = "â­" * (book["rating"] or 0) if book["rating"] else ""
            
            text = f"{icon} {order} {book['title']} {rating}"
            
            item = QListWidgetItem(text)
            item.setData(Qt.ItemDataRole.UserRole, book["id"])
            self.books_list.addItem(item)
    
    def on_series_double_clicked(self, item):
        """Seriye Ã§ift tÄ±klandÄ±ÄŸÄ±nda gÃ¶ster."""
        self.show_series()
    
    def show_series(self):
        """SeÃ§ili seriyi ana listede gÃ¶ster."""
        current = self.series_list.currentItem()
        if current:
            self.selected_series = current.data(Qt.ItemDataRole.UserRole)
            if self.selected_series:
                self.accept()


# ============================================================
# AI ASISTAN DIALOG'U
# ============================================================

class AIWorkerThread(QThread):
    """AI yanÄ±tÄ±nÄ± arka planda alÄ±r."""
    finished = pyqtSignal(str)
    error = pyqtSignal(str)
    
    def __init__(self, func, *args, **kwargs):
        super().__init__()
        self.func = func
        self.args = args
        self.kwargs = kwargs
    
    def run(self):
        try:
            result = self.func(*self.args, **self.kwargs)
            if result:
                self.finished.emit(result)
            else:
                self.error.emit("YanÄ±t alÄ±namadÄ±. Ollama Ã§alÄ±ÅŸÄ±yor mu?")
        except Exception as e:
            self.error.emit(str(e))


class AIAssistantDialog(QDialog):
    """AI Asistan dialog'u - Ollama ile kitap Ã¶nerileri."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        
        self.setWindowTitle("ğŸ¤– AI Kitap AsistanÄ±")
        self.setMinimumSize(700, 600)
        self.setModal(True)
        
        self.worker = None
        self.model = None
        
        self.setup_ui()
        self.check_ollama()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        # Durum Ã§ubuÄŸu
        self.status_frame = QFrame()
        self.status_frame.setStyleSheet("background-color: #2D2D2D; border-radius: 8px; padding: 10px;")
        status_layout = QHBoxLayout(self.status_frame)
        
        self.status_icon = QLabel("â³")
        self.status_icon.setStyleSheet("font-size: 24px;")
        status_layout.addWidget(self.status_icon)
        
        self.status_label = QLabel("Ollama kontrol ediliyor...")
        self.status_label.setStyleSheet("font-size: 14px;")
        status_layout.addWidget(self.status_label, stretch=1)
        
        self.model_combo = QComboBox()
        self.model_combo.setFixedWidth(150)
        self.model_combo.setEnabled(False)
        status_layout.addWidget(self.model_combo)
        
        layout.addWidget(self.status_frame)
        
        # HÄ±zlÄ± eylemler
        actions_label = QLabel("ğŸš€ HÄ±zlÄ± Eylemler:")
        actions_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addWidget(actions_label)
        
        actions_layout = QHBoxLayout()
        
        self.recommend_btn = QPushButton("ğŸ“š Kitap Ã–ner")
        self.recommend_btn.clicked.connect(self.get_recommendations)
        self.recommend_btn.setEnabled(False)
        actions_layout.addWidget(self.recommend_btn)
        
        self.analyze_btn = QPushButton("ğŸ“Š Okuma Analizi")
        self.analyze_btn.clicked.connect(self.analyze_habits)
        self.analyze_btn.setEnabled(False)
        actions_layout.addWidget(self.analyze_btn)
        
        self.plan_btn = QPushButton("ğŸ“… Okuma PlanÄ±")
        self.plan_btn.clicked.connect(self.get_reading_plan)
        self.plan_btn.setEnabled(False)
        actions_layout.addWidget(self.plan_btn)
        
        layout.addLayout(actions_layout)
        
        # Soru sor
        question_layout = QHBoxLayout()
        
        self.question_input = QLineEdit()
        self.question_input.setPlaceholderText("Bir soru sor... (Ã¶rn: 'Dostoyevski gibi baÅŸka yazarlar Ã¶ner')")
        self.question_input.returnPressed.connect(self.ask_question)
        self.question_input.setEnabled(False)
        question_layout.addWidget(self.question_input)
        
        self.ask_btn = QPushButton("Sor")
        self.ask_btn.clicked.connect(self.ask_question)
        self.ask_btn.setEnabled(False)
        question_layout.addWidget(self.ask_btn)
        
        layout.addLayout(question_layout)
        
        # YanÄ±t alanÄ±
        layout.addWidget(QLabel("ğŸ’¬ YanÄ±t:"))
        
        self.response_text = QTextEdit()
        self.response_text.setReadOnly(True)
        self.response_text.setPlaceholderText("AI yanÄ±tÄ± burada gÃ¶rÃ¼necek...")
        self.response_text.setStyleSheet("""
            QTextEdit {
                background-color: #1E1E1E;
                border: 1px solid #3C3C3C;
                border-radius: 8px;
                padding: 10px;
                font-size: 13px;
                line-height: 1.5;
            }
        """)
        layout.addWidget(self.response_text, stretch=1)
        
        # Alt bilgi
        info_label = QLabel("ğŸ’¡ Ä°pucu: Ollama kurulu deÄŸilse 'brew install ollama && ollama pull mistral' ile kurabilirsiniz.")
        info_label.setStyleSheet("color: #888; font-size: 11px;")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        
        # Kapat butonu
        close_btn = QPushButton("Kapat")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)
    
    def check_ollama(self):
        """Ollama durumunu kontrol eder."""
        try:
            from services.ai_service import check_ollama_status
            
            status = check_ollama_status()
            
            if status["available"]:
                self.status_icon.setText("âœ…")
                self.status_label.setText(f"Ollama hazÄ±r!")
                self.model = status["recommended"]
                
                # Modelleri ekle
                self.model_combo.clear()
                for model in status["models"]:
                    self.model_combo.addItem(model)
                
                # Ã–nerilen modeli seÃ§
                index = self.model_combo.findText(self.model)
                if index >= 0:
                    self.model_combo.setCurrentIndex(index)
                
                self.model_combo.setEnabled(True)
                self.model_combo.currentTextChanged.connect(self.on_model_changed)
                
                # ButonlarÄ± aktif et
                self.recommend_btn.setEnabled(True)
                self.analyze_btn.setEnabled(True)
                self.plan_btn.setEnabled(True)
                self.question_input.setEnabled(True)
                self.ask_btn.setEnabled(True)
            else:
                self.status_icon.setText("âŒ")
                self.status_label.setText(status.get("error", "Ollama bulunamadÄ±"))
                
        except ImportError:
            self.status_icon.setText("âŒ")
            self.status_label.setText("AI servisi yÃ¼klenemedi")
        except Exception as e:
            self.status_icon.setText("âŒ")
            self.status_label.setText(f"Hata: {str(e)}")
    
    def on_model_changed(self, model_name):
        """Model deÄŸiÅŸtiÄŸinde."""
        self.model = model_name
    
    def set_loading(self, loading: bool):
        """YÃ¼kleniyor durumunu ayarlar."""
        self.recommend_btn.setEnabled(not loading)
        self.analyze_btn.setEnabled(not loading)
        self.plan_btn.setEnabled(not loading)
        self.ask_btn.setEnabled(not loading)
        self.question_input.setEnabled(not loading)
        
        if loading:
            self.response_text.setPlainText("â³ DÃ¼ÅŸÃ¼nÃ¼yorum...")
            self.status_icon.setText("â³")
        else:
            self.status_icon.setText("âœ…")
    
    def on_response(self, response: str):
        """AI yanÄ±tÄ± geldiÄŸinde."""
        self.set_loading(False)
        self.response_text.setPlainText(response)
    
    def on_error(self, error: str):
        """Hata olduÄŸunda."""
        self.set_loading(False)
        self.response_text.setPlainText(f"âŒ Hata: {error}")
    
    def get_recommendations(self):
        """Kitap Ã¶nerisi al."""
        from services.ai_service import get_book_recommendation
        
        books = db.get_all_books()
        if not books:
            self.response_text.setPlainText("KitaplÄ±ÄŸÄ±nÄ±z boÅŸ. Ã–nce kitap ekleyin!")
            return
        
        self.set_loading(True)
        
        self.worker = AIWorkerThread(
            get_book_recommendation,
            [dict(b) for b in books],
            model=self.model
        )
        self.worker.finished.connect(self.on_response)
        self.worker.error.connect(self.on_error)
        self.worker.start()
    
    def analyze_habits(self):
        """Okuma alÄ±ÅŸkanlÄ±klarÄ±nÄ± analiz et."""
        from services.ai_service import analyze_reading_habits
        
        books = db.get_all_books()
        if not books:
            self.response_text.setPlainText("KitaplÄ±ÄŸÄ±nÄ±z boÅŸ. Ã–nce kitap ekleyin!")
            return
        
        self.set_loading(True)
        
        self.worker = AIWorkerThread(
            analyze_reading_habits,
            [dict(b) for b in books],
            model=self.model
        )
        self.worker.finished.connect(self.on_response)
        self.worker.error.connect(self.on_error)
        self.worker.start()
    
    def get_reading_plan(self):
        """Okuma planÄ± oluÅŸtur."""
        from services.ai_service import get_reading_plan
        
        books = db.get_all_books()
        if not books:
            self.response_text.setPlainText("KitaplÄ±ÄŸÄ±nÄ±z boÅŸ. Ã–nce kitap ekleyin!")
            return
        
        # Hedefi al
        from datetime import datetime
        goal_data = db.get_reading_goal(datetime.now().year)
        goal = goal_data["target_books"] if goal_data else None
        
        self.set_loading(True)
        
        self.worker = AIWorkerThread(
            get_reading_plan,
            [dict(b) for b in books],
            goal=goal,
            model=self.model
        )
        self.worker.finished.connect(self.on_response)
        self.worker.error.connect(self.on_error)
        self.worker.start()
    
    def ask_question(self):
        """Serbest soru sor."""
        from services.ai_service import generate_response, create_book_summary
        
        question = self.question_input.text().strip()
        if not question:
            return
        
        books = db.get_all_books()
        context = create_book_summary([dict(b) for b in books]) if books else "KitaplÄ±k boÅŸ."
        
        prompt = f"""Sen bir kitap uzmanÄ±sÄ±n. KullanÄ±cÄ±nÄ±n kitaplÄ±ÄŸÄ± hakkÄ±nda bilgin var.

{context}

KullanÄ±cÄ±nÄ±n sorusu: {question}

TÃ¼rkÃ§e ve yardÄ±mcÄ± bir ÅŸekilde yanÄ±tla."""
        
        self.set_loading(True)
        self.question_input.clear()
        
        self.worker = AIWorkerThread(
            generate_response,
            prompt,
            model=self.model
        )
        self.worker.finished.connect(self.on_response)
        self.worker.error.connect(self.on_error)
        self.worker.start()


# ============================================================
# YARDIM DIALOG'U
# ============================================================

class HelpDialog(QDialog):
    """YardÄ±m dialog'u - Rehber, kÄ±sayollar, Ã¶zellikler, hakkÄ±nda."""
    
    def __init__(self, parent=None, page: str = "guide"):
        super().__init__(parent)
        
        self.setWindowTitle("ğŸ“š KitaplÄ±ÄŸÄ±m - YardÄ±m")
        self.setMinimumSize(700, 550)
        self.setModal(True)
        
        self.setup_ui()
        self.show_page(page)
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        # BaÅŸlÄ±k
        self.title_label = QLabel()
        self.title_label.setStyleSheet("font-size: 20px; font-weight: bold; padding: 10px;")
        layout.addWidget(self.title_label)
        
        # Ä°Ã§erik
        self.content_text = QTextEdit()
        self.content_text.setReadOnly(True)
        self.content_text.setStyleSheet("""
            QTextEdit {
                background-color: #1E1E1E;
                border: 1px solid #3C3C3C;
                border-radius: 8px;
                padding: 15px;
                font-size: 13px;
                line-height: 1.6;
            }
        """)
        layout.addWidget(self.content_text, stretch=1)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        
        guide_btn = QPushButton("ğŸ“– Rehber")
        guide_btn.clicked.connect(lambda: self.show_page("guide"))
        btn_layout.addWidget(guide_btn)
        
        shortcuts_btn = QPushButton("âŒ¨ï¸ KÄ±sayollar")
        shortcuts_btn.clicked.connect(lambda: self.show_page("shortcuts"))
        btn_layout.addWidget(shortcuts_btn)
        
        features_btn = QPushButton("âœ¨ Ã–zellikler")
        features_btn.clicked.connect(lambda: self.show_page("features"))
        btn_layout.addWidget(features_btn)
        
        about_btn = QPushButton("â„¹ï¸ HakkÄ±nda")
        about_btn.clicked.connect(lambda: self.show_page("about"))
        btn_layout.addWidget(about_btn)
        
        btn_layout.addStretch()
        
        close_btn = QPushButton("Kapat")
        close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)
    
    def show_page(self, page: str):
        """Belirtilen sayfayÄ± gÃ¶sterir."""
        pages = {
            "guide": self.get_guide_content,
            "shortcuts": self.get_shortcuts_content,
            "features": self.get_features_content,
            "about": self.get_about_content,
        }
        
        titles = {
            "guide": "ğŸ“– BaÅŸlangÄ±Ã§ Rehberi",
            "shortcuts": "âŒ¨ï¸ Klavye KÄ±sayollarÄ±",
            "features": "âœ¨ Ã–zellikler",
            "about": "â„¹ï¸ HakkÄ±nda",
        }
        
        self.title_label.setText(titles.get(page, "YardÄ±m"))
        content = pages.get(page, self.get_guide_content)()
        self.content_text.setHtml(content)
    
    def get_guide_content(self) -> str:
        return """
        <h2>ğŸš€ HÄ±zlÄ± BaÅŸlangÄ±Ã§</h2>
        
        <h3>ğŸ“š Kitap Ekleme</h3>
        <p><b>Online Arama ile (Ã–nerilen):</b></p>
        <ol>
            <li><code>Ctrl+N</code> veya <b>Dosya â†’ Ara ve Ekle</b></li>
            <li>Kitap adÄ±, yazar veya ISBN ile arayÄ±n</li>
            <li>SonuÃ§lardan birini seÃ§in</li>
            <li>Bilgileri kontrol edip <b>Kaydet</b></li>
        </ol>
        
        <p><b>Manuel Ekleme:</b></p>
        <ol>
            <li><code>Ctrl+Shift+N</code> veya <b>Dosya â†’ Manuel Ekle</b></li>
            <li>Bilgileri doldurun (sadece baÅŸlÄ±k zorunlu)</li>
            <li><b>ğŸ” Kapak Ara</b> ile kapak bulun</li>
            <li><b>Kaydet</b></li>
        </ol>
        
        <h3>ğŸ“‹ Okuma Listesi KullanÄ±mÄ±</h3>
        <ol>
            <li>KitabÄ±n durumunu <b>ğŸ“‹ OkuyacaÄŸÄ±m</b> yapÄ±n</li>
            <li><b>KitaplÄ±k â†’ Okuma Listesi</b> aÃ§Ä±n (<code>Ctrl+L</code>)</li>
            <li>SÃ¼rÃ¼kle-bÄ±rak ile sÄ±ralayÄ±n</li>
            <li>Okuma hÄ±zÄ±nÄ±zÄ± ayarlayÄ±n</li>
            <li>Tahmini bitiÅŸ tarihini gÃ¶rÃ¼n</li>
        </ol>
        
        <h3>ğŸ—‚ï¸ Raflar</h3>
        <p>Sol panelden:</p>
        <ul>
            <li><b>+ Raf Ekle</b> ile yeni raf oluÅŸturun</li>
            <li>Kitaba saÄŸ tÄ±k â†’ <b>Rafa Ekle</b></li>
            <li>Bir kitap birden fazla rafta olabilir</li>
        </ul>
        
        <h3>ğŸ¤– AI Asistan</h3>
        <ol>
            <li>Ollama'yÄ± kurun: <code>brew install ollama</code></li>
            <li>Model indirin: <code>ollama pull mistral</code></li>
            <li><b>KitaplÄ±k â†’ AI Asistan</b> aÃ§Ä±n (<code>Ctrl+Shift+A</code>)</li>
            <li>Kitap Ã¶nerisi, okuma analizi veya soru sorun</li>
        </ol>
        
        <h3>ğŸ’¡ Ä°puÃ§larÄ±</h3>
        <ul>
            <li><b>Ã‡oklu seÃ§im:</b> <code>Ctrl+Click</code> ile birden fazla kitap seÃ§in</li>
            <li><b>HÄ±zlÄ± dÃ¼zenleme:</b> Tabloda Ã§ift tÄ±k ile dÃ¼zenleyin</li>
            <li><b>Kapak gÃ¶rÃ¼nÃ¼mÃ¼:</b> GÃ¶rÃ¼nÃ¼m â†’ Grid moduna geÃ§in</li>
            <li><b>Seri takibi:</b> Kitap formunda Seri sekmesinden seri bilgisi girin</li>
        </ul>
        """
    
    def get_shortcuts_content(self) -> str:
        return """
        <h2>âŒ¨ï¸ Klavye KÄ±sayollarÄ±</h2>
        
        <h3>ğŸ“š Kitap Ä°ÅŸlemleri</h3>
        <table width="100%" cellpadding="8">
            <tr><td width="40%"><code>Ctrl+N</code></td><td>Online arama ile kitap ekle</td></tr>
            <tr><td><code>Ctrl+Shift+N</code></td><td>Manuel kitap ekle</td></tr>
            <tr><td><code>Delete</code></td><td>SeÃ§ili kitabÄ± sil</td></tr>
            <tr><td><code>Ctrl+Click</code></td><td>Ã‡oklu seÃ§im</td></tr>
            <tr><td><code>Shift+Click</code></td><td>AralÄ±k seÃ§imi</td></tr>
            <tr><td><code>Ã‡ift tÄ±k</code></td><td>KitabÄ± dÃ¼zenle</td></tr>
        </table>
        
        <h3>ğŸ“‹ GÃ¶rÃ¼nÃ¼m</h3>
        <table width="100%" cellpadding="8">
            <tr><td width="40%"><code>Ctrl+L</code></td><td>Okuma listesi</td></tr>
            <tr><td><code>Ctrl+I</code></td><td>Ä°statistikler</td></tr>
            <tr><td><code>Ctrl+B</code></td><td>Kenar Ã§ubuÄŸunu aÃ§/kapat</td></tr>
            <tr><td><code>Ctrl+Shift+A</code></td><td>AI Asistan</td></tr>
        </table>
        
        <h3>ğŸ”§ Genel</h3>
        <table width="100%" cellpadding="8">
            <tr><td width="40%"><code>Ctrl+Q</code></td><td>Uygulamadan Ã§Ä±k</td></tr>
            <tr><td><code>F1</code></td><td>Bu yardÄ±m penceresi</td></tr>
            <tr><td><code>Escape</code></td><td>Dialog'u kapat</td></tr>
        </table>
        
        <h3>ğŸ“ Tabloda Gezinme</h3>
        <table width="100%" cellpadding="8">
            <tr><td width="40%"><code>â†‘ â†“</code></td><td>YukarÄ±/aÅŸaÄŸÄ± git</td></tr>
            <tr><td><code>Page Up/Down</code></td><td>Sayfa sayfa git</td></tr>
            <tr><td><code>Home/End</code></td><td>Ä°lk/son satÄ±ra git</td></tr>
        </table>
        """
    
    def get_features_content(self) -> str:
        return """
        <h2>âœ¨ TÃ¼m Ã–zellikler</h2>
        
        <h3>ğŸ“– Kitap YÃ¶netimi</h3>
        <ul>
            <li><b>Online Arama:</b> Google Books, Open Library, Kitapyurdu, Ä°defix, BKM Kitap</li>
            <li><b>40+ Alan:</b> Ã‡eviri bilgileri, satÄ±n alma, konum, etiketler ve daha fazlasÄ±</li>
            <li><b>Kapak GÃ¶rselleri:</b> Ã‡oklu kaynaktan arama veya dosyadan ekleme</li>
            <li><b>Toplu Ä°ÅŸlemler:</b> Ã‡oklu seÃ§im ile dÃ¼zenleme, silme, rafa ekleme</li>
            <li><b>Kitap Kopyalama:</b> Mevcut kitabÄ± ÅŸablon olarak kullan</li>
            <li><b>AlÄ±ntÄ±lar:</b> Sevilen cÃ¼mleleri sayfa numarasÄ± ile kaydet</li>
        </ul>
        
        <h3>ğŸ“š Kitap Serileri</h3>
        <ul>
            <li>Serileri otomatik grupla</li>
            <li>Seri iÃ§i okuma durumu takibi</li>
            <li>Seri tamamlama yÃ¼zdesi</li>
            <li>Kitap formunda serideki diÄŸer kitaplarÄ± gÃ¶r</li>
        </ul>
        
        <h3>ğŸ“‹ Okuma Listesi</h3>
        <ul>
            <li><b>5 Durum:</b> OkunmadÄ±, OkuyacaÄŸÄ±m, Okunuyor, Okundu, OkumayacaÄŸÄ±m</li>
            <li><b>SÄ±ralama:</b> SÃ¼rÃ¼kle-bÄ±rak ile okuma sÄ±rasÄ±</li>
            <li><b>Tahmini SÃ¼re:</b> GÃ¼nlÃ¼k sayfa hÄ±zÄ±na gÃ¶re hesaplama</li>
            <li><b>BitiÅŸ Tarihi:</b> TÃ¼m listeyi ne zaman bitirirsin</li>
        </ul>
        
        <h3>ğŸ¯ Okuma Takibi</h3>
        <ul>
            <li>Mevcut sayfa takibi</li>
            <li>BaÅŸlama ve bitirme tarihleri</li>
            <li>YÄ±llÄ±k okuma hedefi</li>
            <li>Okuma sayÄ±sÄ±</li>
        </ul>
        
        <h3>ğŸ“Š Ä°statistikler (7 Sekme)</h3>
        <ul>
            <li><b>Genel:</b> Toplam kitap, sayfa, durum daÄŸÄ±lÄ±mÄ±</li>
            <li><b>Yazarlar:</b> En Ã§ok okunan yazarlar</li>
            <li><b>YayÄ±nevleri:</b> YayÄ±nevi daÄŸÄ±lÄ±mÄ±</li>
            <li><b>YÄ±llar:</b> YayÄ±n yÄ±lÄ± analizi</li>
            <li><b>Okuma HÄ±zÄ±:</b> AylÄ±k okuma, en hÄ±zlÄ±/yavaÅŸ kitap</li>
            <li><b>Puanlar:</b> Puan daÄŸÄ±lÄ±mÄ±</li>
            <li><b>Hedef:</b> YÄ±llÄ±k hedef takibi</li>
        </ul>
        
        <h3>ğŸ¤– AI Asistan (Ollama)</h3>
        <ul>
            <li>KiÅŸiselleÅŸtirilmiÅŸ kitap Ã¶nerileri</li>
            <li>Okuma alÄ±ÅŸkanlÄ±klarÄ± analizi</li>
            <li>Okuma planÄ± oluÅŸturma</li>
            <li>Serbest soru sorma</li>
            <li>Tamamen yerel, internet gerektirmez</li>
        </ul>
        
        <h3>ğŸ¨ ArayÃ¼z</h3>
        <ul>
            <li>Koyu ve aÃ§Ä±k tema</li>
            <li>Grid (kapak) ve liste gÃ¶rÃ¼nÃ¼mÃ¼</li>
            <li>Ã–zelleÅŸtirilebilir sÃ¼tunlar</li>
            <li>SatÄ±r iÃ§i dÃ¼zenleme</li>
        </ul>
        
        <h3>ğŸ“¤ Ä°Ã§e/DÄ±ÅŸa Aktarma</h3>
        <ul>
            <li><b>Ä°Ã§e:</b> CSV, Excel</li>
            <li><b>DÄ±ÅŸa:</b> CSV, JSON, Excel</li>
        </ul>
        """
    
    def get_about_content(self) -> str:
        return """
        <div style="text-align: center; padding: 20px;">
            <h1>ğŸ“š KitaplÄ±ÄŸÄ±m</h1>
            <p style="font-size: 16px; color: #888;">Versiyon 1.0</p>
            
            <p style="margin-top: 30px; font-size: 14px;">
                KiÅŸisel kitap koleksiyonunuzu yÃ¶netmek iÃ§in<br>
                modern ve kullanÄ±cÄ± dostu bir masaÃ¼stÃ¼ uygulamasÄ±.
            </p>
            
            <hr style="margin: 30px 0; border-color: #3C3C3C;">
            
            <h3>ğŸ› ï¸ Teknolojiler</h3>
            <p>
                <b>Python 3.10+</b> â€¢ <b>PyQt6</b> â€¢ <b>SQLite</b><br>
                <b>Ollama</b> (AI Asistan iÃ§in)
            </p>
            
            <h3>ğŸ“š Veri KaynaklarÄ±</h3>
            <p>
                Google Books API â€¢ Open Library<br>
                Kitapyurdu â€¢ Ä°defix â€¢ BKM Kitap
            </p>
            
            <hr style="margin: 30px 0; border-color: #3C3C3C;">
            
            <p style="color: #888;">
                MIT LisansÄ± ile lisanslanmÄ±ÅŸtÄ±r.
            </p>
            
            <p style="margin-top: 20px; font-size: 18px;">
                Claude ile â¤ï¸ yapÄ±ldÄ±
            </p>
            
            <p style="color: #888; font-size: 12px; margin-top: 30px;">
                Â© 2024 KitaplÄ±ÄŸÄ±m
            </p>
        </div>
        """